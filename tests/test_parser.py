"""利润宝 · parser.py 单元测试（S1-S5 验证范围）。"""
import os
import sys

import pytest

from core import parser as pr
from core import finance as fin

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)


def test_clean_number_thousands():
    assert pr.clean_number("1,234,567") == 1_234_567.0


def test_clean_number_parentheses_negative():
    assert pr.clean_number("(500)") == -500.0


def test_clean_number_percent_and_currency():
    assert pr.clean_number("¥12,000.50") == 12_000.5


def test_clean_number_none():
    assert pr.clean_number("") is None
    assert pr.clean_number(None) is None


def test_synonym_normalization():
    assert pr.normalize_account_name("主营业务收入") == "营业收入"
    assert pr.normalize_account_name("营业税金及附加") == "税金及附加"
    assert pr.normalize_account_name("营业费用") == "销售费用"


def test_parse_financial_dict_synonyms():
    raw = {
        "company_name": "测试公司",
        "industry": "制造业",
        "years": [2023],
        "income_statement": {
            "主营业务收入": {2023: 1_000_000},
            "营业税金及附加": {2023: 12_000},
        },
        "balance_sheet": {},
        "account_balances": {"招待费": 5_000},
    }
    data = pr.parse_financial_dict(raw)
    assert data.income_statement["营业收入"][2023] == 1_000_000.0
    assert data.income_statement["税金及附加"][2023] == 12_000.0
    assert data.account_balances["业务招待费"] == {2023: 5_000.0}
    assert data.parsed_meta["matched"] >= 3


def test_parse_roundtrip_finance():
    raw = {
        "years": [2023],
        "income_statement": {
            "营业收入": {2023: 1_000_000},
            "营业成本": {2023: 700_000},
            "税金及附加": {2023: 12_000},
            "所得税费用": {2023: 50_000},
            "净利润": {2023: 238_000},
        },
    }
    data = pr.parse_financial_dict(raw)
    ind = fin.compute_year_indicators(data, 2023)
    assert ind["增值税税负率"]["value"] == 10.0
    assert ind["毛利率"]["value"] == 30.0


# ── S5 Excel 解析测试 ────────────────────────────────────────────────────
SAMPLE_XLSX = os.path.join(ROOT, "demo_output", "sample_finance.xlsx")


@pytest.fixture(scope="module")
def ensure_sample_xlsx():
    """确保样例 Excel 存在；不存在则生成。"""
    if not os.path.exists(SAMPLE_XLSX):
        from data import make_sample
        make_sample.write_sample_xlsx(SAMPLE_XLSX)
    return SAMPLE_XLSX


def test_match_sheet_kind():
    assert pr._match_sheet_kind("利润表") == "income"
    assert pr._match_sheet_kind("损益表") == "income"
    assert pr._match_sheet_kind("资产负债表") == "balance"
    assert pr._match_sheet_kind("科目余额表") == "ledger"
    assert pr._match_sheet_kind("余额表") == "ledger"
    assert pr._match_sheet_kind("无关Sheet") is None


def test_reject_xls_raises():
    with pytest.raises(pr.ParserError):
        pr._reject_xls("/tmp/legacy.xls")


def test_parse_excel_three_sheets(ensure_sample_xlsx):
    data = pr.parse_excel(ensure_sample_xlsx, company_name="测试公司", industry="制造业")
    assert data.company_name == "测试公司"
    assert data.years == [2021, 2022, 2023]
    assert "营业收入" in data.income_statement
    assert "资产总额" in data.balance_sheet
    assert "业务招待费" in data.account_balances
    # 三 Sheet 全部识别
    assert set(data.parsed_meta["found_kinds"]) == {"income", "balance", "ledger"}


def test_parse_excel_synonym_merge(ensure_sample_xlsx):
    """样例 Excel 用规范科目名，归并匹配数为 0；但解析结果可计算指标。"""
    data = pr.parse_excel(ensure_sample_xlsx)
    ind = fin.compute_year_indicators(data, 2023)
    # 2023 营收 15,200,000，税金及附加 35,000 → 估算增值税 291,666 → 税负率 1.92%
    assert abs(ind["增值税税负率"]["value"] - 1.92) < 0.01
    assert "估算" in ind["增值税税负率"]["note"]


def test_parse_excel_files_split(ensure_sample_xlsx, tmp_path):
    """分文件导入：用样例 Excel 拆出三个临时文件再合并。"""
    import openpyxl
    src_wb = openpyxl.load_workbook(ensure_sample_xlsx)
    inc_path = tmp_path / "income.xlsx"
    bal_path = tmp_path / "balance.xlsx"
    led_path = tmp_path / "ledger.xlsx"

    for target, sheet_name in [(inc_path, "利润表"), (bal_path, "资产负债表"), (led_path, "科目余额表")]:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        src = src_wb[sheet_name]
        for row in src.iter_rows(values_only=True):
            ws.append(list(row))
        wb.save(str(target))
    src_wb.close()

    data = pr.parse_excel_files(
        income_path=str(inc_path),
        balance_path=str(bal_path),
        ledger_path=str(led_path),
    )
    assert data.years == [2021, 2022, 2023]
    assert "营业收入" in data.income_statement
    assert "资产总额" in data.balance_sheet
    assert "业务招待费" in data.account_balances


def test_parse_excel_missing_table_warning(tmp_path):
    """缺表时 parsed_meta['warnings'] 含明确提示。"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "利润表"
    ws.append(["项目", "2023"])
    ws.append(["营业收入", 1_000_000])
    ws.append(["营业成本", 700_000])
    path = tmp_path / "only_income.xlsx"
    wb.save(str(path))
    wb.close()

    data = pr.parse_excel(str(path))
    warnings_text = "\n".join(data.parsed_meta.get("warnings", []))
    assert "资产负债表" in warnings_text
    assert "科目余额表" in warnings_text


def test_parse_smart_csv_and_excel(ensure_sample_xlsx, tmp_path):
    """parse_smart 自动识别 CSV 与 Excel。"""
    csv_path = tmp_path / "inc.csv"
    with open(csv_path, "w", encoding="utf-8-sig") as fh:
        fh.write("项目,2023\n营业收入,1000000\n营业成本,700000\n")
    data = pr.parse_smart(path=str(csv_path))
    assert data.income_statement["营业收入"][2023] == 1_000_000.0

    data2 = pr.parse_smart(path=ensure_sample_xlsx)
    assert data2.years == [2021, 2022, 2023]


def test_parse_excel_zero_revenue_no_crash(tmp_path):
    """营收为 0 时不崩溃，指标标注基数缺失。"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "利润表"
    ws.append(["项目", "2023"])
    ws.append(["营业收入", 0])
    ws.append(["税金及附加", 12_000])
    path = tmp_path / "zero_rev.xlsx"
    wb.save(str(path))
    wb.close()
    data = pr.parse_excel(str(path))
    ind = fin.compute_year_indicators(data, 2023)
    assert ind["增值税税负率"]["value"] == 0.0
    assert ind["增值税税负率"]["note"] == fin.BASE_MISSING_NOTE
