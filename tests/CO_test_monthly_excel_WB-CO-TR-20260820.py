"""月度执行计划 Sheet 兼容回归。

覆盖：make_sample 生成第一稿 → 追加月度 Sheet → read_template 读回不报错、
Q/R 公式 openpyxl 复算（R 全 0、Q==Σ月）、重复追加幂等。
"""

from __future__ import annotations

import importlib
import os
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

from core import parser as pr
from core import budget_template as tpl_mod
from data import make_sample

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

budget_export = importlib.import_module("core.CO_budget_export_WB-CO-TR-20260810")
monthly = importlib.import_module("core.CO_monthly_split_WB-CO-TR-20260820")


@pytest.fixture
def sample_data():
    return pr.parse_financial_dict(make_sample.build_sample_data())


@pytest.fixture
def draft(tmp_path, sample_data):
    """规则链路生成第一稿（含 plan_rows 快照），返回 (path, meta)。"""
    out = tmp_path / "draft.xlsx"
    path, meta = budget_export.export_budget_3sheet(sample_data, str(out))
    assert Path(path).exists()
    assert meta.get("plan_rows"), "第一稿 meta 应含 plan_rows 快照"
    return str(path), meta


def _split_payload(meta) -> dict:
    result = monthly.rule_split(meta["plan_rows"], {
        "q_season": "下半年旺",
        "q_bonus": "年终奖2~3个月工资",
        "q_campaign": "旺季前1~2个月前置投放",
        "q_lump": "",
    }, budget_year=2023)
    payload = result.to_payload()
    payload["generated_at"] = "2026-08-20 12:00:00"
    return payload


def test_append_then_read_template_ok(draft):
    """AC-A4.2：追加月度 Sheet 后 read_template 读回不报错、原 Sheet 不受影响。"""
    path, meta = draft
    out = budget_export.append_monthly_sheet(path, _split_payload(meta))
    wb = load_workbook(out)
    assert "月度执行计划" in wb.sheetnames
    assert wb.sheetnames[:4] == [
        "费用预算表", "行业企业所得税贡献率参考", "诊断与行动清单", "费用合规筹划约束",
    ]
    # read_template（用户模板读回校验）对追加 Sheet 天然兼容
    plan = tpl_mod.read_template(out)
    assert plan is not None and len(plan.lines) == 84


def test_monthly_formulas_recompute_zero_gap(draft):
    """AC-A4.1：Q=SUM(E:P) 复算 == Σ月；R=Q−D 全 0。"""
    path, meta = draft
    out = budget_export.append_monthly_sheet(path, _split_payload(meta))
    ws = load_workbook(out, data_only=False)["月度执行计划"]
    rows = sorted(meta["plan_rows"], key=lambda r: r["row"])
    for idx, item in enumerate(rows):
        r = 2 + idx
        months = [float(ws.cell(r, 5 + m).value or 0) for m in range(12)]
        q_formula = ws.cell(r, 17).value
        assert str(q_formula).startswith("=SUM(E"), q_formula
        # 引擎口径复算：Σ月 == round(annual)
        assert round(sum(months)) == round(float(item["annual"])), f"第{item['row']}行 Σ月≠年"
        # R 公式存在且为 =Q-D
        assert str(ws.cell(r, 18).value) == f"=Q{r}-D{r}"
    # 用 data_only 打开由 Excel/LibreOffice 计算的缓存不可用（新写公式无缓存），
    # 公式语义由 openpyxl 表达式 + 引擎恒等校验共同保证。


def test_append_idempotent_overwrites(draft):
    """同名 Sheet 重复调用幂等：不重复、以最后一次为准。"""
    path, meta = draft
    payload = _split_payload(meta)
    out1 = budget_export.append_monthly_sheet(path, payload)
    out2 = budget_export.append_monthly_sheet(out1, payload)
    wb = load_workbook(out2)
    assert wb.sheetnames.count("月度执行计划") == 1
    # 行数 = 84 数据行 + 表头 + 合计 + 空行 + 说明
    ws = wb["月度执行计划"]
    assert ws.max_row == 1 + 84 + 2 + 1


def test_download_route_end_to_end(draft):
    """download 路由语义（stage=ready 时可产出含月度 Sheet 文件）。"""
    path, meta = draft
    payload = _split_payload(meta)
    out = budget_export.append_monthly_sheet(path, payload, mode=payload["mode"])
    assert "（含月度拆分）" in Path(out).name
    wb = load_workbook(out)
    ws = wb["月度执行计划"]
    # D 列（年度）与第一稿 G 同源：抽查若干行 annual 一致
    rows = {int(r["row"]): float(r["annual"]) for r in meta["plan_rows"]}
    checked = 0
    for r in range(2, 86):
        row_no = int(ws.cell(r, 1).value or 0)
        if row_no in rows and rows[row_no] > 0:
            assert abs(float(ws.cell(r, 4).value) - rows[row_no]) < 0.01
            checked += 1
    assert checked >= 5, "应至少核到 5 行非零年度金额"
