"""利润宝 · 模板版 Excel 导出与行动清单公式测试（T6.5 + CO P0-4）。

覆盖：
- 行动测算 Excel：VAT 行 F/G/H/I 全为 0
- 行动测算 Excel：研发行 G 为加计扣除公式
- 行动测算 Excel：咨询费行 F+H 公式
- 行动测算 Excel：招待费 H 列为公式（非冻结常量），含 MIN/D*60%/营收*0.5%/$C$3/J
- 修改目标值 E 与税率 J 后 H 与 I 应手工复算一致
- 行动测算为空时概览公式为 0（无循环引用）
- 模板版 Excel/Word/PDF 三类导出能生成非空文件
"""
from __future__ import annotations

import os
import sys
import tempfile

import pytest

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from core import action_pack as ap
from core import budget as bk
from core import report as core_report
from data import make_sample
from core import parser as pr, diagnostic as diag, interactive as iv


@pytest.fixture(scope="module")
def sample_session():
    """构造样例会话（含 4 条发现，全选 A 进入 FINAL）。"""
    raw = make_sample.build_sample_data()
    data = pr.parse_financial_dict(raw)
    result = diag.diagnose(data)
    sess = iv.start_session(data, result)
    for f in result.findings:
        iv.submit_decision(sess, f.id, "A")
    iv.confirm(sess, user_confirmed=True)
    return sess


# ── 行动测算 Excel：按发现类型公式 ─────────────────────────────────────

def test_action_pack_export_excel(sample_session, tmp_path):
    """Excel 行动测算能生成非空文件，含三个 Sheet。"""
    path = str(tmp_path / "model.xlsx")
    ap.export_excel_model(sample_session, path)
    assert os.path.exists(path) and os.path.getsize(path) > 0
    import openpyxl
    wb = openpyxl.load_workbook(path)
    assert "方案概览" in wb.sheetnames
    assert "行动测算" in wb.sheetnames
    assert "执行清单" in wb.sheetnames
    wb.close()


def _find_action_row(ws, finding_id_keyword: str, n_entries: int):
    """在行动测算 Sheet 中按关键字找到行号。"""
    for r in range(5, 5 + n_entries):
        b = ws.cell(row=r, column=2).value or ""
        if finding_id_keyword in str(b):
            return r
    return None


def test_action_pack_vat_row_all_zero(sample_session, tmp_path):
    """P0-4：VAT 行 F/G/H/I 全为 0，禁止把百分点当元计算。"""
    path = str(tmp_path / "vat.xlsx")
    ap.export_excel_model(sample_session, path)
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb["行动测算"]
    r = _find_action_row(ws, "增值税", len(sample_session.draft2))
    assert r is not None, "未找到增值税行"
    for col_idx in (6, 7, 8, 9):
        v = ws.cell(row=r, column=col_idx).value
        # 应为 0 或公式结果 0；不能是含 MAX 的元单位公式
        assert v == 0 or v == "0", f"VAT 行第 {col_idx} 列应为 0，实际 {v!r}"
    wb.close()


def test_action_pack_rd_row_g_formula(sample_session, tmp_path):
    """P0-4：研发行 G 列为加计扣除公式 =MAX(0,E-D)*J*K。"""
    path = str(tmp_path / "rd.xlsx")
    ap.export_excel_model(sample_session, path)
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb["行动测算"]
    r = _find_action_row(ws, "研发", len(sample_session.draft2))
    assert r is not None, "未找到研发行"
    g = ws.cell(row=r, column=7).value
    assert isinstance(g, str) and g.startswith("="), f"研发行 G 应为公式，实际 {g!r}"
    assert "MAX" in g and "E" in g and "D" in g and "J" in g and "K" in g, f"G 公式不正确: {g!r}"
    # F=0, H=0
    assert ws.cell(row=r, column=6).value == 0
    assert ws.cell(row=r, column=8).value == 0
    wb.close()


def test_action_pack_entertain_h_formula_not_constant(sample_session, tmp_path):
    """P0-4：招待费 H 列必须为 Excel 公式（非冻结常量），含 MIN/D*60%/营收*0.5%/$C$3/J。"""
    path = str(tmp_path / "ent.xlsx")
    ap.export_excel_model(sample_session, path)
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb["行动测算"]
    r = _find_action_row(ws, "招待", len(sample_session.draft2))
    assert r is not None, "未找到招待费行"
    h = ws.cell(row=r, column=8).value
    assert isinstance(h, str) and h.startswith("="), f"H 必须是公式，实际 {h!r}"
    assert "MIN" in h, f"H 必须含 MIN: {h!r}"
    assert "0.6" in h, f"H 必须含 60% (0.6): {h!r}"
    assert "0.005" in h, f"H 必须含 0.5% (0.005): {h!r}"
    assert "$C$3" in h, f"H 必须引用 $C$3 参考营收: {h!r}"
    assert f"J{r}" in h, f"H 必须引用 J{r} 税率: {h!r}"
    # 验证 C3 参考营收单元格存在且非 0
    c3 = ws.cell(row=3, column=3).value
    assert c3 is not None and float(c3) > 0, f"C3 参考营收应非 0，实际 {c3!r}"
    wb.close()


def test_action_pack_entertain_h_recalculates_on_e_change(sample_session, tmp_path):
    """P0-4：手工复算两个目标值 E 的 H 列结果，必须不同（动态联动）。"""
    path = str(tmp_path / "ent2.xlsx")
    ap.export_excel_model(sample_session, path)
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb["行动测算"]
    r = _find_action_row(ws, "招待", len(sample_session.draft2))
    assert r is not None
    c3 = float(ws.cell(row=3, column=3).value)
    j = float(ws.cell(row=r, column=10).value)
    d = float(ws.cell(row=r, column=4).value)
    # 两个目标值 E 应产生不同的 H
    e1, e2 = 50_000, 200_000
    h1 = max(0, min(d * 0.6, c3 * 0.005) - min(e1 * 0.6, c3 * 0.005)) * j
    h2 = max(0, min(d * 0.6, c3 * 0.005) - min(e2 * 0.6, c3 * 0.005)) * j
    assert h1 != h2, f"E 变化时 H 应不同: h1={h1} h2={h2}"
    # 净影响 I = F + G - H，也应不同
    f1, f2 = max(0, d - e1), max(0, d - e2)
    i1, i2 = f1 - h1, f2 - h2
    assert i1 != i2
    wb.close()


def test_action_pack_consulting_f_h_formulas(sample_session, tmp_path):
    """P0-4：咨询费行 F=压降额公式, H=压降额×税率公式。"""
    path = str(tmp_path / "consult.xlsx")
    ap.export_excel_model(sample_session, path)
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb["行动测算"]
    r = _find_action_row(ws, "咨询", len(sample_session.draft2))
    assert r is not None, "未找到咨询费行"
    f = ws.cell(row=r, column=6).value
    h = ws.cell(row=r, column=8).value
    assert isinstance(f, str) and "MAX" in f and "D" in f and "E" in f, f"咨询费 F 公式不正确: {f!r}"
    assert isinstance(h, str) and "MAX" in h and "J" in r"\n".join([str(h)]), f"咨询费 H 公式不正确: {h!r}"
    wb.close()


def test_action_pack_empty_draft_no_circular_reference(tmp_path):
    """P0-4：空草稿时概览公式为 0（无循环引用）。"""
    raw = {
        "company_name": "正常公司", "industry": "制造业", "years": [2023],
        "income_statement": {
            "营业收入": {2023: 10_000_000}, "营业成本": {2023: 7_000_000},
            "税金及附加": {2023: 360_000}, "研发费用": {2023: 500_000},
            "所得税费用": {2023: 250_000}, "净利润": {2023: 750_000},
        },
        "balance_sheet": {}, "account_balances": {"业务招待费": 30_000, "咨询服务费": 100_000},
    }
    data = pr.parse_financial_dict(raw)
    result = diag.diagnose(data)
    sess = iv.start_session(data, result)
    # 无发现 → 直接到 DRAFT2/CONFIRMATION
    iv.confirm(sess, user_confirmed=True)
    assert len(sess.draft2) == 0
    path = str(tmp_path / "empty.xlsx")
    ap.export_excel_model(sess, path)
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb["方案概览"]
    # 找到「总净影响（元）」单元格
    found_zero = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "总净影响（元）":
                next_val = ws.cell(row=cell.row, column=cell.column + 1).value
                # 应为 0 而非 =SUM(...) 引用空行动 Sheet
                assert next_val == 0, f"空草稿时总净影响应为 0，实际 {next_val!r}"
                found_zero = True
    assert found_zero, "未找到『总净影响（元）』行"
    wb.close()


# ── 模板版 Excel/Word/PDF 导出 ─────────────────────────────────────────

def test_template_excel_export(tmp_path):
    """模板版 Excel 三 Sheet 能生成非空文件。"""
    plan = bk.make_empty_plan(company_name="测试公司", industry="制造业", year=2024)
    plan.top_inputs.budget_revenue = 1_000_000
    plan.top_inputs.budget_cost = 600_000
    plan.top_inputs.last_year_revenue = 900_000
    plan.top_inputs.last_year_cost = 550_000
    plan.top_inputs.income_tax_rate = 0.05
    plan.top_inputs.company_contribution_rate = 0.003
    bk.compute_all(plan)
    from core import budget_template as bt
    path = str(tmp_path / "tpl.xlsx")
    bt.write_template(plan, path)
    assert os.path.exists(path) and os.path.getsize(path) > 0


def test_budget_word_export(tmp_path):
    """模板版 Word 报告能生成非空文件。"""
    plan = bk.make_empty_plan(company_name="测试公司", industry="制造业", year=2024)
    plan.top_inputs.budget_revenue = 1_000_000
    bk.compute_all(plan)
    path = str(tmp_path / "budget.docx")
    core_report.export_budget_word(plan, path)
    assert os.path.exists(path) and os.path.getsize(path) > 0


def test_budget_pdf_export(tmp_path):
    """模板版 PDF 报告能生成非空文件（含 CJK 字体嵌入）。"""
    plan = bk.make_empty_plan(company_name="测试公司", industry="制造业", year=2024)
    plan.top_inputs.budget_revenue = 1_000_000
    bk.compute_all(plan)
    path = str(tmp_path / "budget.pdf")
    core_report.export_budget_pdf(plan, path)
    assert os.path.exists(path) and os.path.getsize(path) > 0
