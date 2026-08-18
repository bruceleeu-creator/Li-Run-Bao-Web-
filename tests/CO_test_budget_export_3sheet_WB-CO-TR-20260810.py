"""三 Sheet 费用预算模板导出：结构与用户标准模板一致。"""

from __future__ import annotations

import importlib
import os
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

from core import parser as pr
from data import make_sample

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

budget_export = importlib.import_module("core.CO_budget_export_WB-CO-TR-20260810")
# 参考模板优先取环境变量（真实用户模板，不入库）；否则用仓库内示例
_ref_env = os.environ.get("LRB_BUDGET_REF_XLSX", "").strip()
REF = Path(_ref_env) if _ref_env else Path(ROOT) / "demo_output" / "budget_3sheet_WB-CO-TR-20260726.xlsx"


@pytest.fixture
def sample_data():
    return pr.parse_financial_dict(make_sample.build_sample_data())


def test_export_budget_3sheet_matches_template_labels(sample_data, tmp_path):
    out = tmp_path / "out.xlsx"
    path, meta = budget_export.export_budget_3sheet(sample_data, str(out))
    assert Path(path).exists()
    assert meta["sheets"] == [
        "费用预算表",
        "行业企业所得税贡献率参考",
        "诊断与行动清单",
        "费用合规筹划约束",
    ]

    wb = load_workbook(path)
    ref = load_workbook(str(REF))
    # 前三 Sheet 与标准模板一致；第四 Sheet 为合规约束（新增）
    assert wb.sheetnames[:3] == ref.sheetnames == [
        "费用预算表",
        "行业企业所得税贡献率参考",
        "诊断与行动清单",
    ]
    assert "费用合规筹划约束" in wb.sheetnames
    ws_c = wb["费用合规筹划约束"]
    assert "三条" in str(ws_c["A4"].value or "") or "规则" in str(ws_c["A4"].value or "")
    assert "历史" in str(ws_c["A4"].value or "") or any(
        "历史" in str(ws_c.cell(r, 1).value or "") for r in range(1, 40)
    )

    ws, rs = wb["费用预算表"], ref["费用预算表"]
    for coord in ("A1", "A2", "A3", "A4", "A5", "A6", "A12", "A13", "D13", "G13", "I13", "J13"):
        assert ws[coord].value == rs[coord].value, coord
    # 交付建模：占比/毛利预计算写入（打开即见），恒等 E=D/C6、H=G/C2
    assert float(ws["C2"].value) > 0
    c2 = float(ws["C2"].value)
    c6 = float(ws["C6"].value or 0)
    assert isinstance(ws["C4"].value, (int, float))
    assert abs(float(ws["C4"].value) - (c2 - float(ws["C3"].value or 0))) < 0.02
    for r in range(14, 98):
        d = float(ws.cell(r, 4).value or 0)
        e = ws.cell(r, 5).value
        g = float(ws.cell(r, 7).value or 0)
        h = ws.cell(r, 8).value
        j = ws.cell(r, 10).value
        if d > 0 and c6 > 0:
            assert isinstance(e, (int, float)) and float(e) > 0
            assert abs(float(e) - d / c6) < 1e-6
        if g > 0 and c2 > 0:
            assert isinstance(h, (int, float)) and float(h) > 0
            assert abs(float(h) - g / c2) < 1e-6
        if g or float(ws.cell(r, 9).value or 0):
            assert isinstance(j, (int, float))
            assert abs(float(j) - (g - float(ws.cell(r, 9).value or 0))) < 0.02
    # 合计行
    assert isinstance(ws["G98"].value, (int, float))
    assert isinstance(ws["H98"].value, (int, float))


def test_reconcile_plan_to_financials_aligns_subject_di():
    """D/I 必须对齐利润表期间费用合计。"""
    from core import budget as budget_mod
    from core.models import FinancialData

    plan = budget_mod.make_empty_plan(company_name="对账", industry="建筑业", year=2024)
    plan.top_inputs.budget_revenue = 100_000_000
    plan.top_inputs.budget_cost = 80_000_000
    plan.top_inputs.last_year_revenue = 90_000_000
    plan.top_inputs.last_year_cost = 72_000_000
    plan.top_inputs.income_tax_rate = 0.25
    # 财务科目故意错：D 偏小 G 更小
    fin = [l for l in plan.lines if l.subject == "财务费用"]
    fin[0].last_year_actual = 100_000
    fin[0].budget_amount = 50_000
    data = FinancialData(
        company_name="对账",
        industry="建筑业",
        years=[2023, 2024],
        income_statement={
            "营业收入": {2023: 90_000_000, 2024: 100_000_000},
            "营业成本": {2023: 72_000_000, 2024: 80_000_000},
            "销售费用": {2023: 0, 2024: 0},
            "管理费用": {2023: 5_000_000, 2024: 5_500_000},
            "研发费用": {2023: 0, 2024: 0},
            "财务费用": {2023: 2_000_000, 2024: 2_200_000},
        },
        balance_sheet={},
        account_balances={},
    )
    notes = budget_export.reconcile_plan_to_financials(plan, data)
    d_fin = sum(float(l.last_year_actual or 0) for l in plan.lines if l.subject == "财务费用")
    g_fin = sum(float(l.budget_amount or 0) for l in plan.lines if l.subject == "财务费用")
    assert abs(d_fin - 2_000_000) < 1
    assert g_fin > 1_500_000  # 应对齐上年×增长附近
    assert any("财务" in n for n in notes)


def test_export_api_budget_async(sample_data, tmp_path, monkeypatch):
    """异步任务：mock DeepSeek 提取后应 completed 并可下载。"""
    import time
    app_mod = importlib.import_module("web_backend.CO_app_WB-CO-TR-20260805160732")
    session = importlib.import_module("web_backend.CO_session_WB-CO-TR-20260805160732")
    job_mod = importlib.import_module("web_backend.CO_budget_export_job_WB-CO-TR-20260810")
    ai_mod = importlib.import_module("web_backend.CO_ai_WB-CO-TR-20260805160732")
    from fastapi.testclient import TestClient

    out = tmp_path / "exports"
    out.mkdir()
    monkeypatch.setattr(job_mod, "_EXPORT_DIR", out)

    def fake_top(ocr, structured_hints=None):
        h = structured_hints or {}
        return {
            "budget_revenue": float(h.get("budget_revenue") or 1),
            "budget_cost": float(h.get("budget_cost") or 1),
            "last_year_revenue": float(h.get("last_year_revenue") or 1),
            "last_year_cost": float(h.get("last_year_cost") or 1),
            "_period": {},
        }, ""

    def fake_lines(ocr, catalog, structured_facts=None, period_totals=None):
        return (
            [
                {
                    "row": 48,
                    "last_year_actual": 1000,
                    "budget_amount": 1000,
                    "actual_amount": 500,
                }
            ],
            "",
        )

    monkeypatch.setattr(ai_mod, "extract_budget_indicators", fake_top)
    monkeypatch.setattr(ai_mod, "extract_budget_expense_lines", fake_lines)

    client = TestClient(app_mod.create_app())
    session.replace(sample_data, ["OCR 利润表 营业收入"], [])
    r = client.post("/api/export/budget/jobs")
    assert r.status_code == 200, r.text
    jid = r.json()["job_id"]
    job = None
    for _ in range(80):
        job = client.get(f"/api/export/budget/jobs/{jid}").json()
        if job["status"] in ("completed", "failed"):
            break
        time.sleep(0.05)
    assert job and job["status"] == "completed", job
    d = client.get(f"/api/export/budget/jobs/{jid}/download")
    assert d.status_code == 200
    assert len(d.content) > 3000
    client.close()
