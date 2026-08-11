"""利润宝 · Web 财报导入 API 回归测试。

验证导入闭环：上传 xlsx/csv、分文件、示例数据、指标计算、离线回退、
非法输入与 API 契约。全部使用 TestClient，不依赖浏览器。
"""

from __future__ import annotations

import io
import importlib

import pytest
from fastapi.testclient import TestClient

app_mod = importlib.import_module("web_backend.CO_app_WB-CO-TR-20260805160732")
import_mod = importlib.import_module("web_backend.CO_import_WB-CO-TR-20260805160732")
session_mod = importlib.import_module("web_backend.CO_session_WB-CO-TR-20260805160732")

CLIENT = TestClient(app_mod.create_app())


@pytest.fixture(autouse=True)
def _reset_session():
    session_mod.clear()
    yield
    session_mod.clear()


def _sample_xlsx_bytes() -> bytes:
    from data.make_sample import write_sample_xlsx
    from core.parser import ParserError
    import os

    path = write_sample_xlsx()
    try:
        with open(path, "rb") as fh:
            return fh.read()
    finally:
        os.remove(path)


def _xlsx(name: str = "sample.xlsx") -> dict:
    return {"files": ("sample.xlsx", io.BytesIO(_sample_xlsx_bytes()), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}


# ── 单文件 .xlsx ─────────────────────────────────────────────────────────

def test_import_single_xlsx():
    r = CLIENT.post("/api/import", files=_xlsx(), data={"company_name": "样例厂", "industry": "制造业"})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["summary"]["company_name"] == "样例厂"
    assert body["summary"]["industry"] == "制造业"
    assert body["summary"]["years"] == [2021, 2022, 2023]
    assert body["summary"]["latest_year"] == 2023
    assert body["summary"]["matched"] >= 0


def test_saved_previews_persist_after_import_and_clear():
    """导入后预览必须保存到本机会话：切页/刷新后仍可查看，清空会话一并清除。"""
    r = CLIENT.post("/api/import", files=_xlsx(), data={"company_name": "样例厂", "industry": "制造业"})
    assert r.status_code == 200, r.text
    assert r.json()["previews"], "导入响应应包含文件预览"

    saved = CLIENT.get("/api/import/saved-previews")
    assert saved.status_code == 200
    files = saved.json()["files"]
    assert files, "导入后应有已保存预览"
    assert files[0]["name"] == "sample.xlsx"
    assert files[0]["kind"] == "xlsx"

    CLIENT.post("/api/session/clear")
    cleared = CLIENT.get("/api/import/saved-previews").json()["files"]
    assert cleared == [], "清空会话后已保存预览应一并清除"


def test_sample_import_has_no_saved_previews():
    """示例数据没有源文件预览，切换页面不应展示旧导入的预览。"""
    CLIENT.post("/api/import", files=_xlsx(), data={"company_name": "样例厂", "industry": "制造业"})
    CLIENT.post("/api/import/sample")
    assert CLIENT.get("/api/import/saved-previews").json()["files"] == []


def test_import_sample_endpoint():
    r = CLIENT.post("/api/import/sample")
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["summary"]["company_name"] == "示例制造有限公司"
    assert body["summary"]["years"] == [2021, 2022, 2023]
    assert len(body["indicators"]) == 3


def test_sample_indicators_match_finance_core():
    """Web 指标必须与 core.finance 口径一致（Tk 基线与 Web 对照）。"""
    r = CLIENT.post("/api/import/sample")
    ind2023 = r.json()["indicators"][2]

    from data.make_sample import build_sample_data
    from core import parser as p
    from core import finance as f

    data = p.parse_financial_dict(build_sample_data())
    core_ind = f.compute_year_indicators(data, 2023)

    assert ind2023["毛利率"]["value"] == core_ind["毛利率"]["value"]
    assert ind2023["净利率"]["value"] == core_ind["净利率"]["value"]
    assert ind2023["增值税税负率"]["value"] == core_ind["增值税税负率"]["value"]
    assert ind2023["增值税税负率"]["estimate"] is True
    assert "估算值" in ind2023["增值税税负率"]["note"]
    assert ind2023["研发费用率"]["value"] == 0.0


# ── 分文件导入 ───────────────────────────────────────────────────────────

def _build_annual_report_xlsx(year: int, scale: float = 1.0) -> bytes:
    """构造一整年的完整审计报告 xlsx（利润表+资产负债表+科目余额表三 Sheet）。

    金额按 scale 缩放，使三年数据可区分；科目余额表仅反映该年期末余额。
    """
    import openpyxl

    from data.make_sample import INCOME, BALANCE, LEDGER

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "利润表"
    ws1.append(["项目", str(year)])
    for acc in INCOME:
        ws1.append([acc, int(INCOME[acc][year] * scale)])
    ws2 = wb.create_sheet("资产负债表")
    ws2.append(["项目", str(year)])
    for acc in BALANCE:
        ws2.append([acc, int(BALANCE[acc][year] * scale)])
    ws3 = wb.create_sheet("科目余额表")
    ws3.append(["科目名称", "期末余额"])
    for acc, val in LEDGER.items():
        ws3.append([acc, int(val * scale)])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def test_import_split_files():
    """旧「分表文件」语义已废弃：多文件现按「每年完整报告」合并。

    3 个单年度完整年报合并后，科目余额表按年保留。
    """
    files = [
        ("files", ("2021年审计报告.xlsx", io.BytesIO(_build_annual_report_xlsx(2021)), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ("files", ("2022年审计报告.xlsx", io.BytesIO(_build_annual_report_xlsx(2022)), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ("files", ("2023年审计报告.xlsx", io.BytesIO(_build_annual_report_xlsx(2023)), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
    ]
    r = CLIENT.post("/api/import", files=files, data={"company_name": "三年厂", "industry": "制造业"})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["summary"]["years"] == [2021, 2022, 2023]
    # 科目余额表按年保留
    data = session_mod.get_data()
    assert data is not None
    assert data.account_balances.get("业务招待费") == {2021: 260_000.0, 2022: 260_000.0, 2023: 260_000.0}


# ── CSV 导入 ─────────────────────────────────────────────────────────────

def test_import_csv():
    csv_text = "项目,2021,2022,2023\n营业收入,12000000,13500000,15200000\n营业成本,9600000,10530000,11930000\n"
    r = CLIENT.post("/api/import", files={"files": ("利润表.csv", io.BytesIO(csv_text.encode()), "text/csv")})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["summary"]["years"] == [2021, 2022, 2023]


# ── 会话与非法输入 ───────────────────────────────────────────────────────

def test_session_before_import_is_empty():
    r = CLIENT.get("/api/session")
    assert r.status_code == 200
    assert r.json()["session"] is None


def test_session_after_import():
    CLIENT.post("/api/import/sample")
    r = CLIENT.get("/api/session")
    body = r.json()
    assert body["session"]["company_name"] == "示例制造有限公司"
    assert len(body["indicators"]) == 3


def test_clear_session_resets():
    """重新导入：清空会话后返回空，后端数据清空。"""
    CLIENT.post("/api/import/sample")
    assert session_mod.get_data() is not None
    r = CLIENT.post("/api/session/clear")
    assert r.status_code == 200
    body = r.json()
    assert body["session"] is None
    assert body["indicators"] == []
    # 后端 session 数据确实清空
    assert session_mod.get_data() is None
    assert session_mod.get_ocr_texts() == []


def test_session_persists_and_restores_from_db():
    """导入后数据写入 SQLite，清内存后可 restore 恢复（刷新/重启不丢）。"""
    db_mod = importlib.import_module("web_backend.CO_db_WB-CO-TR-20260805160732")
    # 用真实样例数据导入（写穿 DB）
    CLIENT.post("/api/import/sample")
    data = session_mod.get_data()
    assert data is not None
    assert session_mod.get_ocr_texts() == []  # 样例无 OCR
    # 模拟刷新：清内存，DB 里应有数据
    session_mod._data = None
    session_mod._ocr_texts = []
    loaded = db_mod.load_session()
    assert loaded is not None
    assert loaded["company_name"] == "示例制造有限公司"
    # restore_from_db 应恢复内存
    session_mod.restore_from_db()
    restored = session_mod.get_data()
    assert restored is not None
    assert restored.company_name == "示例制造有限公司"
    assert restored.income_statement.get("营业收入", {}).get(2023) == 15_200_000


def test_import_rejects_bad_extension():
    r = CLIENT.post("/api/import", files={"files": ("report.txt", io.BytesIO(b"hi"), "text/plain")})
    assert r.status_code == 400
    assert "仅支持" in r.json()["detail"]


def test_import_empty_files():
    r = CLIENT.post("/api/import", files=[])
    # FastAPI 对空 files 直接 422（缺必需参数）；业务层对空列表 400。两者都拒绝。
    assert r.status_code in (400, 422)


def test_import_three_annual_reports_merges_years():
    """三个完整年报合并成三年数据集。"""
    files = [
        ("files", ("2021年审计报告.xlsx", io.BytesIO(_build_annual_report_xlsx(2021)), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ("files", ("2022年审计报告.xlsx", io.BytesIO(_build_annual_report_xlsx(2022, scale=1.2)), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ("files", ("2023年审计报告.xlsx", io.BytesIO(_build_annual_report_xlsx(2023, scale=1.5)), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
    ]
    r = CLIENT.post("/api/import", files=files, data={"company_name": "三年厂", "industry": "制造业"})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["summary"]["years"] == [2021, 2022, 2023]
    assert len(body["indicators"]) == 3
    # 科目余额表保留三年（各年期末余额不同）
    data = session_mod.get_data()
    assert data is not None
    balances = data.account_balances.get("业务招待费", {})
    assert balances[2021] == 260_000
    assert balances[2022] == int(260_000 * 1.2)
    assert balances[2023] == int(260_000 * 1.5)


def test_industries_lists_benchmarks():
    r = CLIENT.get("/api/industries")
    assert r.status_code == 200
    body = r.json()
    # 新契约：industries 为 [{name, desc}]，names 为旧式名称列表（向后兼容）
    assert "制造业" in body["names"]
    assert body["default"] == "制造业"
    items = {item["name"]: item for item in body["industries"]}
    assert "制造业" in items
    assert items["制造业"]["desc"]
    assert len(body["names"]) >= 10  # 行业已扩展到 10 个


def test_industries_recommend_rule_path():
    """未配置 AI 时行业推荐走规则路径并返回可应用结果。"""
    r = CLIENT.post(
        "/api/industries/recommend",
        json={"company_name": "某某软件科技有限公司", "overview": ""},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["industry"] == "软件和信息技术服务业"
    assert body["source"] == "rule"
    assert "理由" in body["reason"] or body["reason"]
    assert body["fallback"] is False


def test_industries_recommend_fallback_default():
    """无关键词时回退默认行业。"""
    r = CLIENT.post(
        "/api/industries/recommend",
        json={"company_name": "未知名字", "overview": ""},
    )
    assert r.status_code == 200
    assert r.json()["industry"] == "制造业"


# ── 模板工作台预算计划 ──────────────────────────────────────────────────

def test_budget_from_session_requires_import():
    """未导入财报时生成预算计划应提示先导入。"""
    r = CLIENT.post("/api/budget/from-session")
    assert r.status_code == 400
    assert "导入" in r.json()["detail"]


def test_budget_from_session_extracts_indicators():
    """从示例数据生成预算计划：本年/上年营收与成本映射到 TopInputs。

    示例数据无 OCR 文本 → 回退结构化提取（method=fallback）。
    """
    CLIENT.post("/api/import/sample")
    r = CLIENT.post("/api/budget/from-session")
    assert r.status_code == 200, r.text
    body = r.json()
    ti = body["plan"]["top_inputs"]
    # 样例数据：2023 营收 1520万，2022 营收 1350万；成本 1193万 / 1053万
    assert body["plan"]["year"] == 2023
    assert ti["budget_revenue"] == 15_200_000
    assert ti["budget_cost"] == 11_930_000
    assert ti["last_year_revenue"] == 13_500_000
    assert ti["last_year_cost"] == 10_530_000
    assert "结构化提取" in body["source_note"]
    assert body["method"] == "fallback"
    # 比例字段保持模板默认（E4 高新 15%）
    assert ti["income_tax_rate"] == 0.15


def test_budget_from_session_uses_ai(monkeypatch):
    """有 OCR 文本且 AI 可用时，用 AI 识别结果（method=ai）。"""
    ai_mod = importlib.import_module("web_backend.CO_ai_WB-CO-TR-20260805160732")

    class FakeEngine:
        def is_available(self):
            return True

        def chat(self, *args, **kwargs):
            return (
                '{"budget_revenue": 222000000, "budget_cost": 190000000,'
                ' "last_year_revenue": 390000000, "last_year_cost": 350000000}'
            )

    monkeypatch.setattr(ai_mod, "_engine", lambda **kw: FakeEngine())
    # 载入示例数据（无 OCR），再手动塞 OCR 文本触发 AI 分支
    CLIENT.post("/api/import/sample")
    session_mod.set_ocr_texts(["[第 5 页 OCR] 利润表 营业收入 222000000"])
    r = CLIENT.post("/api/budget/from-session")
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["method"] == "ai"
    assert "AI 识别" in body["source_note"]
    ti = body["plan"]["top_inputs"]
    assert ti["budget_revenue"] == 222000000.0
    assert ti["last_year_revenue"] == 390000000.0
