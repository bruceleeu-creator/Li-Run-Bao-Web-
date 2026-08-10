"""利润宝 · 示例数据生成（S4 + S5 扩展）。

生成可复现的制造业样例（含典型问题：研发费用缺失、业务招待费超限、
咨询服务费偏高、增值税税负率偏低）：
- 写入 data/sample_data.json（确定性 JSON，纯标准库）
- 写入 demo_output/sample_finance.xlsx（三 Sheet Excel，含既定异常埋点）
确定性输出（无随机数）。Excel 生成依赖 openpyxl（ADR-004/007）。
"""
from __future__ import annotations

import json
import os
import sys

# 项目根目录加入路径（以脚本方式运行时保证 `import core` 可用）
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

SAMPLE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sample_data.json")
DEMO_DIR = os.path.join(ROOT, "demo_output")
SAMPLE_XLSX_PATH = os.path.join(DEMO_DIR, "sample_finance.xlsx")

COMPANY = "示例制造有限公司"
INDUSTRY = "制造业"
YEARS = [2021, 2022, 2023]

# 利润表（单位：元）—— 确定的样例数值，刻意构造典型问题
INCOME = {
    "营业收入":   {2021: 12_000_000, 2022: 13_500_000, 2023: 15_200_000},
    "营业成本":   {2021: 9_600_000,  2022: 10_530_000, 2023: 11_930_000},
    "税金及附加": {2021: 28_000,     2022: 31_000,     2023: 35_000},   # 偏低 → 增值税税负率异常
    "销售费用":   {2021: 720_000,    2022: 810_000,    2023: 912_000},
    "管理费用":   {2021: 1_320_000,  2022: 1_480_000,  2023: 1_670_000},
    "研发费用":   {2021: 0,          2022: 0,          2023: 0},        # 缺失 → 该有没的
    "财务费用":   {2021: 240_000,    2022: 250_000,    2023: 260_000},
    "利润总额":   {2021: 240_000,    2022: 399_000,    2023: 493_000},
    "所得税费用": {2021: 60_000,     2022: 99_750,     2023: 123_250},
    "净利润":     {2021: 180_000,    2022: 299_250,    2023: 369_750},
}
BALANCE = {
    "资产总额":   {2021: 18_000_000, 2022: 19_500_000, 2023: 21_000_000},
    "负债总额":   {2021: 8_000_000,  2022: 8_600_000,  2023: 9_200_000},
    "所有者权益": {2021: 10_000_000, 2022: 10_900_000, 2023: 11_800_000},
    "应收账款":   {2021: 3_000_000,  2022: 3_400_000,  2023: 4_100_000},
    "存货":       {2021: 2_500_000,  2022: 2_800_000,  2023: 3_200_000},
    "固定资产":   {2021: 6_000_000,  2022: 6_300_000,  2023: 6_600_000},
}
LEDGER = {
    "福利费": 180_000,
    "教育经费": 20_000,
    "业务招待费": 260_000,    # 超限（> 营收 0.5% @2023 ≈ 7.6 万）→ 应控
    "广告宣传费": 400_000,
    "咨询服务费": 980_000,    # 偏高 → 真实性关注
    "研发费用": 0,
    "折旧": 600_000,
}


def build_sample_data() -> dict:
    """返回原始 dict（可被 parser.parse_financial_dict 消费）。"""
    return {
        "company_name": COMPANY,
        "industry": INDUSTRY,
        "years": YEARS,
        "income_statement": INCOME,
        "balance_sheet": BALANCE,
        "account_balances": LEDGER,
    }


def write_sample(path: str = SAMPLE_PATH) -> str:
    data = build_sample_data()
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=2)
    return path


def write_sample_xlsx(path: str = SAMPLE_XLSX_PATH) -> str:
    """生成三 Sheet 样例 Excel（demo_output/sample_finance.xlsx）。

    Sheet1 利润表、Sheet2 资产负债表、Sheet3 科目余额表；
    刻意保留研发费用缺失、业务招待费超限、咨询服务费偏高、增值税税负率偏低等异常埋点。
    """
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError(
            "未安装 openpyxl，无法生成样例 Excel。请运行："
            "python3 -m pip install openpyxl"
        ) from e

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = openpyxl.Workbook()

    # Sheet1：利润表（首行表头：项目 + 年份）
    ws1 = wb.active
    ws1.title = "利润表"
    ws1.append(["项目"] + [str(y) for y in YEARS])
    for acc in INCOME.keys():
        ws1.append([acc] + [INCOME[acc][y] for y in YEARS])

    # Sheet2：资产负债表
    ws2 = wb.create_sheet("资产负债表")
    ws2.append(["项目"] + [str(y) for y in YEARS])
    for acc in BALANCE.keys():
        ws2.append([acc] + [BALANCE[acc][y] for y in YEARS])

    # Sheet3：科目余额表（科目名称 + 期末余额）
    ws3 = wb.create_sheet("科目余额表")
    ws3.append(["科目名称", "期末余额"])
    for acc, val in LEDGER.items():
        ws3.append([acc, val])

    wb.save(path)
    return path


def main() -> int:
    write_sample()
    print(f"[make_sample] 已生成样例 JSON：{SAMPLE_PATH}")
    try:
        xlsx_path = write_sample_xlsx()
        print(f"[make_sample] 已生成样例 Excel：{xlsx_path}")
    except Exception as e:
        print(f"[make_sample] Excel 生成失败：{e}")
        return 1
    print(f"  企业：{COMPANY}（{INDUSTRY}）")
    print(f"  年份：{YEARS}")
    print(f"  典型问题：研发费用缺失 / 业务招待费超限 / 咨询服务费偏高 / 增值税税负率偏低")

    # 简要核验：JSON 与 Excel 双向解析一致（离线，无需大模型）
    from core import parser as p
    from core import finance as f

    fd_json = p.parse_financial_dict(build_sample_data())
    print(f"  [JSON] 同义词归并匹配：{fd_json.parsed_meta.get('matched', 0)} 项")

    fd_xlsx = p.parse_excel(xlsx_path, company_name=COMPANY, industry=INDUSTRY)
    found = fd_xlsx.parsed_meta.get("found_kinds", [])
    print(f"  [Excel] 识别 Sheet 类型：{found}；年份：{fd_xlsx.years}")
    print(f"  [Excel] 同义词归并匹配：{fd_xlsx.parsed_meta.get('matched', 0)} 项")
    if fd_xlsx.parsed_meta.get("warnings"):
        print(f"  [Excel] 警告：{fd_xlsx.parsed_meta['warnings']}")

    for yr in YEARS:
        ind = f.compute_year_indicators(fd_xlsx, yr)
        vat = ind["增值税税负率"]
        print(f"  {yr} 营收={ind['营业收入']:,.0f} 增值税税负率={vat['value']}% [{vat['note']}] "
              f"毛利率={ind['毛利率']['value']}%")
    return 0


if __name__ == "__main__":
    sys.exit(main())
