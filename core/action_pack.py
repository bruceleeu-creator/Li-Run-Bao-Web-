"""利润宝 · Excel 测算模型导出（S6）。

生成 `demo_output/sample_model.xlsx`，含三个 Sheet：
1. 方案概览：企业 / 行业 / 年度 / 落地性 / 总预计节税 / 合规声明 / 估算口径标注
2. 行动测算：每条决策行 = 发现→选项→当前值→目标值→变动幅度→适用税率→预计节税
   目标值与税率列可编辑；变动幅度与预计节税为公式联动（修改目标值即重算）
3. 执行清单：逐条行动 + 状态下拉（待启动/进行中/已完成/已搁置）+ 备注

依赖 openpyxl（ADR-004/007）；导出失败抛 ActionPackError，调用方提示用户。
所有建议限于合法税务筹划；增值税税负率口径须显著标注。
"""
from __future__ import annotations

import os
from typing import List

from .diagnostic import COMPLIANCE_NOTE
from .interactive import Draft2Entry, Session
from . import finance as fin


class ActionPackError(Exception):
    """Excel 测算模型生成错误。"""


def _fmt_money(v: float) -> str:
    return f"{v:,.0f} 元"


# ── 样式 ────────────────────────────────────────────────────────────────
_HEADER_FILL = "1E40AF"
_HEADER_FONT_COLOR = "FFFFFF"
_NOTE_FONT_COLOR = "B91C1C"


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        return openpyxl, Alignment, Border, Font, PatternFill, Side, get_column_letter, DataValidation
    except ImportError as e:
        raise ActionPackError(
            "未安装 openpyxl，无法生成 Excel 测算模型。"
            "请运行：python3 -m pip install openpyxl"
        ) from e


def _apply_header_style(cell, Font, PatternFill, Alignment):
    cell.font = Font(color=_HEADER_FONT_COLOR, bold=True, size=11)
    cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_note_style(cell, Font, Alignment):
    cell.font = Font(color=_NOTE_FONT_COLOR, bold=True, size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _apply_border(cell, Border, Side):
    side = Side(style="thin", color="BFBFBF")
    cell.border = Border(left=side, right=side, top=side, bottom=side)


def _build_overview_sheet(wb, openpyxl, Font, PatternFill, Alignment, Border, Side, sess: Session) -> None:
    """Sheet1：方案概览（关键指标改为公式联动行动测算 Sheet，禁止硬编码汇总）。"""
    ws = wb.active
    ws.title = "方案概览"

    ws["A1"] = "利润宝 · 企业财税优化方案测算模型"
    ws["A1"].font = Font(bold=True, size=14, color="1E3A8A")
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # 静态元数据（不参与计算）
    rows = [
        ("企业名称", sess.data.company_name),
        ("所属行业", sess.data.industry + ("（未匹配，已回退制造业基准）" if sess.diagnosis.industry_fallback else "")),
        ("分析年度", "、".join(str(y) for y in sess.data.years)),
        ("诊断发现数", str(len(sess.diagnosis.findings))),
        ("已决策数", str(len(sess.decisions))),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True, size=11)
        ws.cell(row=i, column=2, value=v).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
        for c in range(1, 5):
            _apply_border(ws.cell(row=i, column=c), Border, Side)

    # 关键指标：公式联动「行动测算」Sheet（修改目标值/税率时自动重算）
    # CO 第二轮复核：空草稿时不得引用行动 Sheet 合计行（循环引用），改为常量 0
    n_entries = len(sess.draft2)
    action_first = 5
    action_last = action_first + max(0, n_entries - 1)
    if n_entries > 0:
        formula_rows = [
            # (label, value, number_format)
            ("总成本节约（元）", f"=SUM(行动测算!F{action_first}:F{action_last})", '#,##0.00'),
            ("总税收节约（元）", f"=SUM(行动测算!G{action_first}:G{action_last})", '#,##0.00'),
            ("总税负影响（元）", f"=SUM(行动测算!H{action_first}:H{action_last})", '#,##0.00'),
            ("总净影响（元）", f"=SUM(行动测算!I{action_first}:I{action_last})", '#,##0.00'),
        ]
    else:
        # 无条目：直接置 0，不引用行动 Sheet（避免循环引用）
        formula_rows = [
            ("总成本节约（元）", 0, '#,##0.00'),
            ("总税收节约（元）", 0, '#,##0.00'),
            ("总税负影响（元）", 0, '#,##0.00'),
            ("总净影响（元）", 0, '#,##0.00'),
        ]
    formula_rows += [
        ("落地性评分（%）", round(sess.feasibility_score, 2), '0.00'),
        ("当前状态", sess.state, '@'),
    ]
    base = 3 + len(rows)
    for i, (label, val, fmt) in enumerate(formula_rows):
        r = base + i
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=11)
        cell = ws.cell(row=r, column=2, value=val)
        cell.number_format = fmt
        cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        for c in range(1, 5):
            _apply_border(ws.cell(row=r, column=c), Border, Side)
        # 公式联动行加重点
        if isinstance(val, str) and val.startswith("="):
            cell.font = Font(bold=True, size=11, color="B91C1C")

    # 落地性扣分明细
    base_row = base + len(formula_rows) + 1
    ws.cell(row=base_row, column=1, value="落地性扣分明细").font = Font(bold=True, size=11, color="1E40AF")
    if sess.feasibility_breakdown:
        for j, b in enumerate(sess.feasibility_breakdown, start=1):
            ws.cell(row=base_row + j, column=1, value=f"  - {b}")
            ws.merge_cells(start_row=base_row + j, start_column=1, end_row=base_row + j, end_column=4)
    else:
        ws.cell(row=base_row + 1, column=1, value="  - 无扣分（默认全选 A 或无可扣分项）")
        ws.merge_cells(start_row=base_row + 1, start_column=1, end_row=base_row + 1, end_column=4)

    # 合规声明
    note_row = base_row + len(sess.feasibility_breakdown) + 3
    ws.cell(row=note_row, column=1, value="合规声明").font = Font(bold=True, size=11, color="B91C1C")
    ws.cell(row=note_row + 1, column=1, value=COMPLIANCE_NOTE)
    ws.merge_cells(start_row=note_row + 1, start_column=1, end_row=note_row + 1, end_column=4)
    _apply_note_style(ws.cell(row=note_row + 1, column=1), Font, Alignment)
    ws.row_dimensions[note_row + 1].height = 60

    # 增值税估算口径显著标注
    ws.cell(row=note_row + 3, column=1, value="增值税税负率估算口径").font = Font(bold=True, size=11, color="B91C1C")
    ws.cell(row=note_row + 4, column=1, value=fin.VAT_ESTIMATE_NOTE)
    ws.merge_cells(start_row=note_row + 4, start_column=1, end_row=note_row + 4, end_column=4)
    _apply_note_style(ws.cell(row=note_row + 4, column=1), Font, Alignment)
    ws.cell(row=note_row + 5, column=1, value=(
        "公式：估算增值税 = 税金及附加 ÷ 12%（增值税附加税费占增值税比例的经验值）；"
        "税负率 = 估算增值税 ÷ 营业收入 × 100%。实际以企业增值税申报表为准。"
    ))
    ws.merge_cells(start_row=note_row + 5, start_column=1, end_row=note_row + 5, end_column=4)
    _apply_note_style(ws.cell(row=note_row + 5, column=1), Font, Alignment)
    ws.row_dimensions[note_row + 5].height = 45

    # 战略意图
    if sess.strategy_notes:
        sn_row = note_row + 7
        ws.cell(row=sn_row, column=1, value="战略意图记录").font = Font(bold=True, size=11, color="1E40AF")
        for k, n in enumerate(sess.strategy_notes, start=1):
            ws.cell(row=sn_row + k, column=1, value=f"  - {n}")
            ws.merge_cells(start_row=sn_row + k, start_column=1, end_row=sn_row + k, end_column=4)

    # 列宽
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 24


def _build_action_calc_sheet(wb, openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter, sess: Session) -> None:
    """Sheet2：行动测算（按发现类型区分公式，CO 第二轮复核要求）。

    列结构（11 列）：
    A 序号 / B 发现 / C 选项 / D 当前值 / E 目标值（可编辑）/
    F 成本节约 / G 税收节约 / H 税负影响 / I 净影响 /
    J 适用税率（可编辑，项目测算默认值/待核验）/ K 加计扣除比例（可编辑，待核验）

    顶部参考输入（C3，可编辑·待核验）：
    - 参考营业收入（元）：用于招待费双重限额公式 MIN(发生额×60%, 营收×0.5%)
      默认取诊断会话最新年度营业收入；用户可改写以核验不同情景

    按发现类型公式规则（CO 第三轮复核：H 必须为公式联动，禁止冻结常量）：
    - VAT_LOW（增值税，单位 %）：F/G/H/I 全为 0（禁止把百分点当元计算）
    - RD_MISSING（研发，金额增加）：F=0、H=0，G=加计扣除节税公式
    - ENTERTAIN_EXCESS（招待费）：
        F = MAX(0, D-E)                              成本节约=压降额
        G = 0                                         招待费无加计扣除
        H = MAX(0, MIN(D*60%,$C$3*0.5%) - MIN(E*60%,$C$3*0.5%)) * J
                                                       双重限额可扣减少额 × 税率
        I = F+G-H                                     净影响
    - CONSULTING_HIGH / 其他（咨询费等全额可扣）：F=压降额、G=0、H=压降额×税率
    """
    ws = wb.create_sheet("行动测算")

    # 说明区
    ws["A1"] = "行动测算表（目标值/适用税率/加计扣除比例/参考营业收入可编辑；其余为公式联动）"
    ws["A1"].font = Font(bold=True, size=12, color="1E3A8A")
    ws.merge_cells("A1:K1")
    ws["A2"] = (
        "注：增值税税负率为估算口径，不直接节税；净影响 = 成本节约 + 税收节约 - 税负影响（单位均为元）。"
        "适用税率、加计扣除比例与参考营业收入均为项目测算默认值，须以企业实际适用政策核验。"
    )
    ws["A2"].font = Font(color=_NOTE_FONT_COLOR, italic=True, size=10)
    ws.merge_cells("A2:K2")

    # 顶部参考输入：参考营业收入（C3，可编辑·待核验）
    # 用于招待费双重限额公式；默认取诊断会话最新年度营业收入
    ws["A3"] = "参考营业收入（元，可编辑·待核验）："
    ws["A3"].font = Font(bold=True, size=10, color="1E40AF")
    ws["A3"].alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells("A3:B3")
    default_revenue = 0.0
    try:
        latest_year = sess.data.latest_year() if sess.data else None
        if latest_year is not None:
            ind = fin.compute_year_indicators(sess.data, latest_year)
            default_revenue = float(ind.get("营业收入", 0.0) or 0.0)
    except Exception:
        default_revenue = 0.0
    revenue_cell = ws.cell(row=3, column=3, value=round(default_revenue, 2))
    revenue_cell.font = Font(bold=True, size=10, color="B91C1C")
    revenue_cell.number_format = '#,##0.00'
    revenue_cell.fill = PatternFill("solid", fgColor="FEF3C7")
    revenue_cell.alignment = Alignment(horizontal="right", vertical="center")
    _apply_border(revenue_cell, Border, Side)
    ws["D3"] = "← 修改此单元格可重算所有招待费行的 H 列（双重限额可扣减少额×税率）"
    ws["D3"].font = Font(color=_NOTE_FONT_COLOR, italic=True, size=9)
    ws.merge_cells("D3:K3")

    # 表头
    headers = [
        "序号", "发现", "选项", "当前值", "目标值（可编辑）",
        "成本节约", "税收节约", "税负影响", "净影响",
        "适用税率（可编辑·待核验）", "加计扣除比例（可编辑·待核验）",
    ]
    header_row = 4
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=ci, value=h)
        _apply_header_style(cell, Font, PatternFill, Alignment)
        _apply_border(cell, Border, Side)
    ws.row_dimensions[header_row].height = 36

    # 数据行 + 公式联动（按发现类型区分）
    entries: List[Draft2Entry] = sess.draft2
    for i, e in enumerate(entries, start=1):
        r = header_row + i
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=e.finding_title)
        ws.cell(row=r, column=3, value=f"{e.option_label}. {e.option_name}")
        ws.cell(row=r, column=4, value=round(e.current_value, 2))
        ws.cell(row=r, column=5, value=round(e.target_value, 2))                # 可编辑

        fid = e.finding_id
        # CO 第二轮复核：按发现类型写公式，禁止统一 MAX(0,D-E)
        if fid == "VAT_LOW":
            # 增值税：价外税不直接节税，且单位为 %，F/G/H/I 全为 0
            ws.cell(row=r, column=6, value=0)
            ws.cell(row=r, column=7, value=0)
            ws.cell(row=r, column=8, value=0)
            ws.cell(row=r, column=9, value=0)
        elif fid == "RD_MISSING":
            # 研发：金额增加，成本节约=0、税负影响=0，税收节约=加计扣除
            ws.cell(row=r, column=6, value=0)
            ws.cell(row=r, column=7, value=f"=MAX(0,E{r}-D{r})*J{r}*K{r}")
            ws.cell(row=r, column=8, value=0)
            ws.cell(row=r, column=9, value=f"=F{r}+G{r}-H{r}")
        elif fid == "ENTERTAIN_EXCESS":
            # CO 第三轮复核：H 必须为公式联动，禁止冻结常量
            # 双重限额：扣除前 = MIN(D*60%, 营收*0.5%)；扣除后 = MIN(E*60%, 营收*0.5%)
            # H = MAX(0, 扣除前 - 扣除后) * J
            # 营收取 C3（顶部参考输入，可编辑）；E 修改后 H 与 I 自动重算
            ws.cell(row=r, column=6, value=f"=MAX(0,D{r}-E{r})")
            ws.cell(row=r, column=7, value=0)
            ws.cell(
                row=r, column=8,
                value=f"=MAX(0,MIN(D{r}*0.6,$C$3*0.005)-MIN(E{r}*0.6,$C$3*0.005))*J{r}",
            )
            ws.cell(row=r, column=9, value=f"=F{r}+G{r}-H{r}")
        else:
            # 咨询费等全额可扣项：F=压降额、G=0、H=压降额×税率
            ws.cell(row=r, column=6, value=f"=MAX(0,D{r}-E{r})")
            ws.cell(row=r, column=7, value=0)
            ws.cell(row=r, column=8, value=f"=MAX(0,D{r}-E{r})*J{r}")
            ws.cell(row=r, column=9, value=f"=F{r}+G{r}-H{r}")

        # 适用税率：使用 Draft2Entry.tax_rate（来自 Option，已标注「项目测算默认值/待核验」）
        ws.cell(row=r, column=10, value=round(float(e.tax_rate), 4))
        # 加计扣除比例：使用 Draft2Entry.deduction_rate（研发=1.0，其他=0.0；待核验）
        ws.cell(row=r, column=11, value=round(float(e.deduction_rate), 4))
        for ci in range(1, 12):
            cell = ws.cell(row=r, column=ci)
            _apply_border(cell, Border, Side)
            cell.alignment = Alignment(horizontal="left" if ci in (2, 3) else "right",
                                       vertical="center", wrap_text=True)
        # 数字格式
        for ci in (4, 5, 6, 7, 8, 9):
            ws.cell(row=r, column=ci).number_format = '#,##0.00'
        ws.cell(row=r, column=10).number_format = '0.00%'
        ws.cell(row=r, column=11).number_format = '0.00'

    # 合计行
    total_row = header_row + len(entries) + 1
    ws.cell(row=total_row, column=2, value="合计").font = Font(bold=True, size=11)
    if entries:
        ws.cell(row=total_row, column=6, value=f"=SUM(F{header_row + 1}:F{header_row + len(entries)})")
        ws.cell(row=total_row, column=7, value=f"=SUM(G{header_row + 1}:G{header_row + len(entries)})")
        ws.cell(row=total_row, column=8, value=f"=SUM(H{header_row + 1}:H{header_row + len(entries)})")
        ws.cell(row=total_row, column=9, value=f"=SUM(I{header_row + 1}:I{header_row + len(entries)})")
    else:
        for ci in (6, 7, 8, 9):
            ws.cell(row=total_row, column=ci, value=0)
    for ci in (6, 7, 8, 9):
        cell = ws.cell(row=total_row, column=ci)
        cell.font = Font(bold=True, size=11, color="B91C1C")
        cell.number_format = '#,##0.00'
    for ci in range(1, 12):
        _apply_border(ws.cell(row=total_row, column=ci), Border, Side)

    # 列宽
    widths = [6, 26, 30, 14, 16, 16, 16, 16, 16, 14, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 冻结表头
    ws.freeze_panes = "A5"


def _build_checklist_sheet(wb, openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter, DataValidation, sess: Session) -> None:
    """Sheet3：执行清单。状态下拉（待启动/进行中/已完成/已搁置）+ 备注。"""
    ws = wb.create_sheet("执行清单")

    ws["A1"] = "执行清单（逐条跟踪）"
    ws["A1"].font = Font(bold=True, size=12, color="1E3A8A")
    ws.merge_cells("A1:F1")

    headers = ["序号", "发现", "行动选项", "操作细节", "状态", "备注"]
    header_row = 3
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=ci, value=h)
        _apply_header_style(cell, Font, PatternFill, Alignment)
        _apply_border(cell, Border, Side)
    ws.row_dimensions[header_row].height = 28

    entries = sess.draft2
    for i, e in enumerate(entries, start=1):
        r = header_row + i
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=e.finding_title)
        ws.cell(row=r, column=3, value=f"{e.option_label}. {e.option_name}")
        ws.cell(row=r, column=4, value=e.action_detail)
        ws.cell(row=r, column=5, value="待启动")  # 默认状态
        ws.cell(row=r, column=6, value="")        # 备注，空
        for ci in range(1, 7):
            cell = ws.cell(row=r, column=ci)
            _apply_border(cell, Border, Side)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 48

    # 状态下拉（数据验证）
    if entries:
        dv = DataValidation(
            type="list",
            formula1='"待启动,进行中,已完成,已搁置"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.error = "请选择有效状态"
        dv.errorTitle = "无效状态"
        dv.prompt = "点击单元格右侧下拉选择状态"
        dv.promptTitle = "状态选择"
        ws.add_data_validation(dv)
        first = header_row + 1
        last = header_row + len(entries)
        dv.add(f"E{first}:E{last}")

    # 合规提示
    tip_row = header_row + len(entries) + 2
    ws.cell(row=tip_row, column=1, value=COMPLIANCE_NOTE)
    ws.merge_cells(start_row=tip_row, start_column=1, end_row=tip_row, end_column=6)
    _apply_note_style(ws.cell(row=tip_row, column=1), Font, Alignment)
    ws.row_dimensions[tip_row].height = 50

    # 列宽
    widths = [6, 26, 28, 50, 12, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"


def export_excel_model(sess: Session, path: str) -> str:
    """生成 Excel 测算模型（含三 Sheet）。

    返回路径；失败抛 ActionPackError。调用方应保留内存方案并提示用户。
    """
    (openpyxl, Alignment, Border, Font, PatternFill, Side,
     get_column_letter, DataValidation) = _require_openpyxl()

    try:
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        wb = openpyxl.Workbook()
        _build_overview_sheet(wb, openpyxl, Font, PatternFill, Alignment, Border, Side, sess)
        _build_action_calc_sheet(wb, openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter, sess)
        _build_checklist_sheet(wb, openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter, DataValidation, sess)
        wb.save(path)
        return path
    except ActionPackError:
        raise
    except Exception as e:
        raise ActionPackError(f"Excel 测算模型生成失败：{type(e).__name__}: {e}") from e
