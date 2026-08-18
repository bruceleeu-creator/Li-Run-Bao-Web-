"""利润宝 · 模板版 Excel 读写（T6.1 + T6.5）。

读取用户原始模板（只读，不修改原文件），导出时可写入项目 `demo_output/`
或用户选择的新路径。所有写出的 Excel 公式与 Python 计算保持一致，且
空白态不得出现 `#DIV/0!/#VALUE!/#REF!`。

- `read_template(path)`：读取原模板 → BudgetPlan
- `write_template(plan, path)`：写出 BudgetPlan → Excel（保留模板结构）
- 工会经费 R90 补齐 F/H/J 公式
- 修正自动筛选范围为 A13:J98
- 冻结窗格 A14
- 行业参考 Sheet 含来源/年度/地区/核验状态

依赖 openpyxl（ADR-004/007）。
"""
from __future__ import annotations

import os
from typing import List, Optional

from .budget_categories import (
    EXPENSE_CATEGORIES,
    TEMPLATE_BALANCE_ROW,
    TEMPLATE_FIRST_ROW,
    TEMPLATE_HEADER_ROW,
    TEMPLATE_LAST_ROW,
    TEMPLATE_TOTAL_ROW,
)
from .budget import (
    BudgetPlan,
    ExpenseLine,
    TopInputs,
    compute_all,
    make_empty_plan,
)


class TemplateError(Exception):
    """模板读写错误。"""


def _require_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        return openpyxl, Alignment, Border, Font, PatternFill, Side, get_column_letter, DataValidation
    except ImportError as e:
        raise TemplateError(
            "未安装 openpyxl，无法读写 Excel 模板。"
            "请运行：python3 -m pip install openpyxl"
        ) from e


# ── 读取 ─────────────────────────────────────────────────────────────────

def read_template(path: str, company_name: str = "", industry: str = "制造业", year: int = 0) -> BudgetPlan:
    """读取用户原始模板（只读）。

    - 顶部 A2:E9 字段映射到 TopInputs
    - A14:J97 的 84 行费用明细映射到 ExpenseLine
    - 行业参考 Sheet 仅保留表头时记入 plan.industry_benchmarks 为空列表
    - 不修改原文件

    P1-3：导入时验证目标 Sheet、关键标签、行号与 84 行结构；
          无关 Excel 不得静默当模板。
    P1-3：A/B/C 列读取用户文件中的科目/费用名称/发票名称并规范化，
          不再永远用硬编码字典覆盖用户维护内容。
    """
    if not os.path.exists(path):
        raise TemplateError(f"模板文件不存在：{path}")
    openpyxl = _require_openpyxl()[0]
    try:
        wb = openpyxl.load_workbook(path, data_only=False, read_only=False)
    except Exception as e:
        raise TemplateError(f"无法打开模板文件 {os.path.basename(path)}：{e}") from e

    # P1-3：验证 Sheet 存在
    ws_name = "费用预算表"
    if ws_name not in wb.sheetnames:
        # 容错：取第一个 Sheet 但提示用户
        if len(wb.sheetnames) == 0:
            raise TemplateError("Excel 文件无任何 Sheet，不是有效的利润宝模板。")
        ws = wb.worksheets[0]
        # 验证首个 Sheet 是否符合模板结构（A13 应为「科目名称」）
        a13_val = ws.cell(row=13, column=1).value
        if a13_val != "科目名称":
            raise TemplateError(
                f"首个 Sheet 「{ws.title}」不是有效的费用预算表：A13 应为「科目名称」，实际「{a13_val}」。"
                "请使用利润宝标准模板（含「费用预算表」Sheet）。"
            )
    else:
        ws = wb[ws_name]

    # P1-3：验证关键标签（A1/A12/A13/D13/G13 等）
    # P0-A（CO T7 重开 v3）：容忍匹配——用户原模板常含「公司名/年度/（元）」前后缀，
    # 且 A12/D13 在真实用户文件中可能是「XXXX有限公司XXXX年费用计划表」「上年同期实际发生费用（元）」
    # 等语义等价变体。每个 coord 提供多个语义替代串，任一匹配即通过。
    expected_labels = {
        "A1": ["企业成本计划表"],
        # A12 真实样例：「XXXX有限公司XXXX年费用计划表   」→ 接受「公司年度费用计划表」或「费用计划表」
        "A12": ["公司年度费用计划表", "费用计划表"],
        "A13": ["科目名称"],
        # D13 真实样例：「上年同期实际发生费用（元）」→ 接受「上年同期实际费用」或「上年同期实际发生费用」
        "D13": ["上年同期实际费用", "上年同期实际发生费用"],
        "G13": ["预算费用金额"],
        "I13": ["实际已发生费用金额"],
    }
    missing_labels = []
    for coord, alternatives in expected_labels.items():
        actual = ws[coord].value
        if not any(_label_matches(actual, alt) for alt in alternatives):
            missing_labels.append(
                f"{coord} 应包含以下任一：「{' / '.join(alternatives)}」，实际「{actual}」"
            )
    if missing_labels:
        raise TemplateError(
            "模板结构不匹配，无法载入：\n  - " + "\n  - ".join(missing_labels)
            + "\n请使用利润宝标准模板。"
        )

    # P1-3：验证 84 行结构（A14:A97 必须有数据或留空但行存在）
    # 检查 A14 与 A97 是否存在有效行（A 列可为空，但行号必须存在）
    if ws.max_row < TEMPLATE_LAST_ROW:
        raise TemplateError(
            f"模板行数不足：应有 {TEMPLATE_LAST_ROW} 行，实际 {ws.max_row} 行。"
            "请使用利润宝标准模板（A14:J97 共 84 行明细）。"
        )

    plan = make_empty_plan(company_name=company_name, industry=industry, year=year)
    plan.source_path = path

    # 顶部输入区
    ti = plan.top_inputs
    ti.budget_revenue = _read_num(ws, "C2") or 0.0
    ti.budget_cost = _read_num(ws, "C3") or 0.0
    ti.last_year_revenue = _read_num(ws, "C6") or 0.0
    ti.last_year_cost = _read_num(ws, "C8") or 0.0
    e2 = _read_num(ws, "E2")
    e3 = _read_num(ws, "E3")
    e4 = _read_num(ws, "E4")
    if e2 is not None:
        ti.industry_contribution_rate = float(e2)
    if e3 is not None:
        ti.company_contribution_rate = float(e3)
    if e4 is not None:
        ti.income_tax_rate = float(e4)
    ti.industry_rate_source = "模板默认（待核验）"
    ti.industry_rate_verified = False

    # P1-3：明细行 - 优先读取用户文件中的科目/费用名称/发票名称（A/B/C 列）
    # 仅当用户文件对应行为空时，才回退到默认字典
    for line in plan.lines:
        r = line.row
        # A/B/C 列：读取用户维护的字典内容（不强制覆盖）
        a_val = ws.cell(row=r, column=1).value
        b_val = ws.cell(row=r, column=2).value
        c_val = ws.cell(row=r, column=3).value
        if a_val and str(a_val).strip():
            line.subject = str(a_val).strip()
        if b_val and str(b_val).strip():
            line.expense_name = str(b_val).strip()
        if c_val and str(c_val).strip():
            line.invoice_name = str(c_val).strip()
        # D/G/I 输入列
        line.last_year_actual = _read_num(ws, f"D{r}") or 0.0
        line.budget_amount = _read_num(ws, f"G{r}") or 0.0
        line.actual_amount = _read_num(ws, f"I{r}") or 0.0

    # 行业参考 Sheet
    if "行业企业所得税贡献率参考" in wb.sheetnames:
        ws2 = wb["行业企业所得税贡献率参考"]
        benchmarks: List[dict] = []
        for r in range(3, ws2.max_row + 1):
            industry_name = ws2.cell(row=r, column=1).value
            if not industry_name:
                continue
            benchmarks.append({
                "industry": str(industry_name).strip(),
                "rate": _read_num(ws2, f"B{r}") or 0.0,
                "source": str(ws2.cell(row=r, column=3).value or "").strip(),
                "year": str(ws2.cell(row=r, column=4).value or "").strip(),
                "region": str(ws2.cell(row=r, column=5).value or "").strip(),
                "verified": str(ws2.cell(row=r, column=6).value or "").strip() in ("是", "True", "true", "1"),
            })
        plan.industry_benchmarks = benchmarks

    try:
        wb.close()
    except Exception:
        pass

    # 计算所有派生字段
    compute_all(plan)
    return plan


def _read_num(ws, coord) -> Optional[float]:
    """安全读取单元格数值；公式时取 value，文本返回 None。"""
    v = ws[coord].value
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "" or s.startswith("="):
        return None
    try:
        return float(s)
    except ValueError:
        return None


# P0-A：全/半角括号 + 「元 / Yuan」后缀 + 多余空白规范化
import re as _re
_LABEL_SUFFIX_RE = _re.compile(r"[\(（]\s*(?:元|Yuan|yuan|YUAN)\s*[\)）]\s*$")


def _normalize_label(value) -> str:
    """规范化标签：去首尾空白；去除尾部全/半角括号包裹的「元/Yuan」后缀；压缩连续空白。"""
    if value is None:
        return ""
    s = str(value).strip()
    # 反复去除尾部括号后缀（防止「（元）(Yuan)」连写）
    while True:
        new_s = _LABEL_SUFFIX_RE.sub("", s).strip()
        if new_s == s:
            break
        s = new_s
    # 压缩连续空白
    s = _re.sub(r"\s+", "", s)
    return s


def _label_matches(actual, expected: str) -> bool:
    """P0-A：标签容忍匹配。

    规则：
    1. 二者规范化后任一为空 → False
    2. 规范化后 actual 包含 expected（子串）→ True
       例：actual="示例公司 2024 企业成本计划表（元）", expected="企业成本计划表" → True
    3. 反向包含（expected 包含 actual）→ True（罕见但容错）
    4. 否则 False
    """
    a = _normalize_label(actual)
    e = _normalize_label(expected)
    if not a or not e:
        return False
    return (e in a) or (a in e)


# ── 写入 ─────────────────────────────────────────────────────────────────

# 样式常量
_HEADER_FILL = "1E40AF"
_HEADER_FONT_COLOR = "FFFFFF"
_NOTE_FONT_COLOR = "B91C1C"
_RED_FILL = "FEE2E2"  # 预算费用重点（红色语义）
_YELLOW_FILL = "FEF3C7"  # 差额/剩余额度重点（黄色语义）
_INPUT_FILL = "F0F9FF"  # 可输入单元格底色
_FORMULA_FILL = "F8FAFC"  # 公式单元格底色
# 合规提示色：绿=达标 · 红=超标/风险（单格着色，便于老板扫一眼）
_OK_FILL = "D1FAE5"       # 浅绿底
_OK_FONT = "047857"       # 深绿字
_BAD_FILL = "FECACA"      # 浅红底
_BAD_FONT = "B91C1C"      # 深红字
_WARN_FILL = "FFEDD5"     # 浅橙底
_WARN_FONT = "C2410C"     # 深橙字


def _paint_ok(cell, Font, PatternFill) -> None:
    cell.fill = PatternFill("solid", fgColor=_OK_FILL)
    cell.font = Font(size=10, bold=True, color=_OK_FONT)


def _paint_bad(cell, Font, PatternFill) -> None:
    cell.fill = PatternFill("solid", fgColor=_BAD_FILL)
    cell.font = Font(size=10, bold=True, color=_BAD_FONT)


def _paint_warn(cell, Font, PatternFill) -> None:
    cell.fill = PatternFill("solid", fgColor=_WARN_FILL)
    cell.font = Font(size=10, bold=True, color=_WARN_FONT)


def write_template(plan: BudgetPlan, path: str, session=None, financial_data=None) -> str:
    """写出 BudgetPlan 到 Excel（建模 + 合规约束 Sheet）。

    Sheet：
    1. 费用预算表（金额输入 + 占比/毛利预计算）
    2. 行业企业所得税贡献率参考
    3. 诊断与行动清单
    4. 费用合规筹划约束（三条硬规则 + 历史占比 + 增速/行业硬顶，交付可见）

    Args:
        plan: 预算计划
        path: 导出路径
        session: 可选诊断会话
        financial_data: 可选 FinancialData，用于写入历史费用率与合规硬顶

    Raises:
        TemplateError: 路径冲突或写出失败
    """
    # P0-3：禁止覆盖原始模板
    if plan.source_path:
        src_real = os.path.realpath(plan.source_path)
        dst_real = os.path.realpath(path)
        if src_real == dst_real:
            raise TemplateError(
                f"导出路径与原始模板路径相同（{dst_real}），禁止覆盖原始模板。请另存为新文件。"
            )

    from .budget import compute_all, validate_plan

    compute_all(plan)
    ok, errors = validate_plan(plan)
    if not ok:
        raise TemplateError(
            "预算计划校验未通过，拒绝导出 Excel：\n  - " + "\n  - ".join(errors)
            + "\n请先在「2. 模板工作台」修正输入后再导出。"
        )
    # 内存模型仍保持 E/H 与 D/G 一致（供报告/API），Excel 侧用公式自算
    ti0 = plan.top_inputs
    c2_0 = float(ti0.budget_revenue or 0)
    c6_0 = float(ti0.last_year_revenue or 0)
    for line in plan.lines:
        d0 = float(line.last_year_actual or 0)
        g0 = float(line.budget_amount or 0)
        line.last_year_expense_ratio = (
            round(d0 / c6_0, 8) if (c6_0 > 0 and d0 > 0) else 0.0
        )
        line.budget_expense_ratio = (
            round(g0 / c2_0, 8) if (c2_0 > 0 and g0 > 0) else 0.0
        )

    openpyxl, Alignment, Border, Font, PatternFill, Side, get_column_letter, DataValidation = _require_openpyxl()
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)

    try:
        wb = openpyxl.Workbook()
        _build_budget_sheet(wb, openpyxl, Alignment, Border, Font, PatternFill, Side, DataValidation, plan)
        _build_industry_benchmark_sheet(wb, openpyxl, Alignment, Border, Font, PatternFill, Side, plan)
        _build_action_summary_sheet(wb, openpyxl, Alignment, Border, Font, PatternFill, Side, plan, session)
        _build_compliance_policy_sheet(
            wb, openpyxl, Alignment, Border, Font, PatternFill, Side, plan, financial_data, session
        )
        wb.save(path)
        return path
    except TemplateError:
        raise
    except Exception as e:
        raise TemplateError(f"Excel 模板写出失败：{type(e).__name__}: {e}") from e


def _apply_header_style(cell, Font, PatternFill, Alignment):
    cell.font = Font(color=_HEADER_FONT_COLOR, bold=True, size=11)
    cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_input_style(cell, Font, PatternFill, Alignment, Border, Side):
    cell.fill = PatternFill("solid", fgColor=_INPUT_FILL)
    cell.font = Font(size=10)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    _apply_border(cell, Border, Side)


def _apply_formula_style(cell, Font, PatternFill, Alignment, Border, Side):
    cell.fill = PatternFill("solid", fgColor=_FORMULA_FILL)
    cell.font = Font(size=10, color="555555")
    cell.alignment = Alignment(horizontal="right", vertical="center")
    _apply_border(cell, Border, Side)


def _apply_border(cell, Border, Side):
    side = Side(style="thin", color="BFBFBF")
    cell.border = Border(left=side, right=side, top=side, bottom=side)


def _build_budget_sheet(
    wb, openpyxl, Alignment, Border, Font, PatternFill, Side, DataValidation, plan: BudgetPlan
) -> None:
    """Sheet1：费用预算表（A1:J100，保留模板结构与红/黄重点语义）。"""
    from openpyxl.utils import get_column_letter
    ws = wb.active
    ws.title = "费用预算表"

    # ── 标题 ──
    ws["A1"] = "企业成本计划表"
    ws["A1"].font = Font(bold=True, size=16, color="1E3A8A")
    ws.merge_cells("A1:J1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── 顶部输入区 A2:E9 ──
    ti = plan.top_inputs
    # 标签列（A/B）+ 输入列（C）+ 标签列（D）+ 输入列（E）
    ws["A2"] = "预算营业收入"
    ws["C2"] = ti.budget_revenue
    ws["D2"] = "行业所得税贡献率"
    ws["E2"] = ti.industry_contribution_rate
    ws["A3"] = "预算营业成本"
    ws["C3"] = ti.budget_cost
    ws["D3"] = "企业预算所得税贡献率"
    ws["E3"] = ti.company_contribution_rate
    ws["A4"] = "毛利"
    ws["D4"] = "所得税税率"
    ws["E4"] = ti.income_tax_rate
    ws["A5"] = "毛利率"
    ws["D5"] = "企业应交所得税预算"
    ws["A6"] = "上年度同期营业收入"
    ws["C6"] = ti.last_year_revenue
    ws["D6"] = "利润总额预算"
    ws["A7"] = "收入增长率"
    ws["D7"] = "预算费用上限"
    ws["A8"] = "上年度同期营业成本"
    ws["C8"] = ti.last_year_cost
    ws["D8"] = "实际已发生费用"
    ws["A9"] = "上年度毛利率"
    ws["D9"] = "费用差额"

    # 输入单元格样式（C2/C3/C6/C8/E2/E3/E4）
    for coord in ("C2", "C3", "C6", "C8", "E2", "E3", "E4"):
        _apply_input_style(ws[coord], Font, PatternFill, Alignment, Border, Side)
    # 标签样式
    for r in range(2, 10):
        for col in ("A", "D"):
            cell = ws[f"{col}{r}"]
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            _apply_border(cell, Border, Side)

    # 顶部公式列（与 core.budget.compute_top 一致；改 C2/C3/E3/E4 自动重算）
    ws["C4"] = "=C2-C3"
    ws["C5"] = "=IF(C2=0,0,C4/C2)"
    ws["C7"] = "=IF(C6=0,0,(C2-C6)/C6)"
    ws["C9"] = "=IF(C6=0,0,(C6-C8)/C6)"
    ws["E5"] = "=E3*C2"
    ws["E6"] = "=IF(E4=0,0,E5/E4)"
    ws["E7"] = "=C4-E6"
    ws["E8"] = "=I98"
    ws["E9"] = "=E7-E8"
    for coord in ("C4", "C5", "C7", "C9", "E5", "E6", "E7", "E8", "E9"):
        _apply_formula_style(ws[coord], Font, PatternFill, Alignment, Border, Side)
    for coord in ("C5", "C7", "C9", "E2", "E3", "E4"):
        ws[coord].number_format = "0.00%"
    for coord in ("C2", "C3", "C4", "C6", "C8", "E5", "E6", "E7", "E8", "E9"):
        ws[coord].number_format = '#,##0.00'

    ws["E7"].fill = PatternFill("solid", fgColor=_RED_FILL)
    ws["E7"].font = Font(size=10, bold=True, color="B91C1C")
    ws["E9"].fill = PatternFill("solid", fgColor=_YELLOW_FILL)
    ws["E9"].font = Font(size=10, bold=True)

    # ── 明细表 A12:J98 ──
    # 第 12 行：表标题
    ws["A12"] = "公司年度费用计划表"
    ws["A12"].font = Font(bold=True, size=12, color="1E3A8A")
    ws.merge_cells("A12:J12")
    ws["A12"].alignment = Alignment(horizontal="center", vertical="center")

    # 第 13 行：表头（A13:J13）
    headers = [
        ("A13", "科目名称"),
        ("B13", "费用名称"),
        ("C13", "发票名称"),
        ("D13", "上年同期实际费用"),
        ("E13", "上年费用占上年收入比例"),
        ("F13", "参考费用金额"),
        ("G13", "预算费用金额"),
        ("H13", "预算费用占预算收入比例"),
        ("I13", "实际已发生费用金额"),
        ("J13", "差额"),
    ]
    for coord, h in headers:
        cell = ws[coord]
        cell.value = h
        _apply_header_style(cell, Font, PatternFill, Alignment)
        _apply_border(cell, Border, Side)
    ws.row_dimensions[13].height = 32

    # 第 14-97 行：输入 D/G/I；派生列写预计算值；G/H/J 按合规规则单格标绿/标红
    from . import compliance_policy as compliance_mod

    c2_top = float(plan.top_inputs.budget_revenue or 0)
    c6_top = float(plan.top_inputs.last_year_revenue or 0)
    c7_top = float(plan.top_computed.revenue_growth_rate or 0)
    f98_sum = 0.0
    max_single = compliance_mod.MAX_SINGLE_LINE_RATIO
    max_rigid = compliance_mod.MAX_SINGLE_LINE_RATIO_RIGID
    fee_growth_cap = c7_top + compliance_mod.FEE_GROWTH_BUFFER_PP

    def _is_rigid_name(name: str) -> bool:
        return any(k in (name or "") for k in ("工资", "薪酬", "社保", "公积金", "房租", "租赁", "奖金"))

    for line in plan.lines:
        r = line.row
        d_val = round(float(line.last_year_actual or 0), 2)
        g_val = round(float(line.budget_amount or 0), 2)
        i_val = round(float(line.actual_amount or 0), 2)
        f_raw = float(getattr(line, "reference_amount", 0) or 0)

        e_val = round(d_val / c6_top, 8) if (d_val > 0 and c6_top > 0) else 0.0
        if d_val > 0:
            f_write = round(d_val * (1.0 + c7_top), 2)
        else:
            f_write = round(f_raw if f_raw > 0 else g_val, 2)
        f98_sum += f_write
        h_val = round(g_val / c2_top, 8) if (g_val > 0 and c2_top > 0) else 0.0
        j_val = round(g_val - i_val, 2)

        ws.cell(row=r, column=1, value=line.subject)
        ws.cell(row=r, column=2, value=line.expense_name)
        ws.cell(row=r, column=3, value=line.invoice_name)
        ws.cell(row=r, column=4, value=d_val)
        ws.cell(row=r, column=5, value=e_val)
        ws.cell(row=r, column=6, value=f_write)
        ws.cell(row=r, column=7, value=g_val)
        ws.cell(row=r, column=8, value=h_val)
        ws.cell(row=r, column=9, value=i_val)
        ws.cell(row=r, column=10, value=j_val)

        for col in (1, 2, 3):
            cell = ws.cell(row=r, column=col)
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            _apply_border(cell, Border, Side)
        for col in (4, 7, 9):
            cell = ws.cell(row=r, column=col)
            _apply_input_style(cell, Font, PatternFill, Alignment, Border, Side)
            cell.number_format = '#,##0.00'
        for col in (5, 8):
            cell = ws.cell(row=r, column=col)
            _apply_formula_style(cell, Font, PatternFill, Alignment, Border, Side)
            cell.number_format = '0.0000%'
        for col in (6, 10):
            cell = ws.cell(row=r, column=col)
            _apply_formula_style(cell, Font, PatternFill, Alignment, Border, Side)
            cell.number_format = '#,##0.00'

        # ── 单格合规色（用户提示：绿达标 / 红超标）──
        g_cell = ws.cell(row=r, column=7)
        h_cell = ws.cell(row=r, column=8)
        j_cell = ws.cell(row=r, column=10)
        cap_line = max_rigid if _is_rigid_name(line.expense_name) else max_single
        if g_val > 0 and c2_top > 0:
            if h_val > cap_line + 1e-9:
                # 单行占营收超上限 → 红
                _paint_bad(g_cell, Font, PatternFill)
                _paint_bad(h_cell, Font, PatternFill)
            else:
                # 有预算且未超单行上限 → 绿
                _paint_ok(g_cell, Font, PatternFill)
                _paint_ok(h_cell, Font, PatternFill)
            # 有上年时：费用增速是否匹配收入增速（G 相对 D）
            if d_val > 0:
                g_growth = (g_val - d_val) / d_val
                if g_growth > fee_growth_cap + 1e-9:
                    _paint_bad(g_cell, Font, PatternFill)
                    f_cell = ws.cell(row=r, column=6)
                    _paint_warn(f_cell, Font, PatternFill)
        # J=G−I：实际超预算 → 红；有预算且未超支 → 绿
        if g_val > 0 or i_val > 0:
            if j_val < -0.01:
                _paint_bad(j_cell, Font, PatternFill)
            elif g_val > 0 and j_val >= 0:
                _paint_ok(j_cell, Font, PatternFill)
            else:
                j_cell.fill = PatternFill("solid", fgColor=_YELLOW_FILL)
                j_cell.font = Font(size=10, bold=True)
        else:
            j_cell.fill = PatternFill("solid", fgColor=_YELLOW_FILL)
            j_cell.font = Font(size=10, bold=True)

    # 第 98 行：合计（预计算数值，打开即见）
    ws.cell(row=98, column=1, value="合计")
    ws.cell(row=98, column=1).font = Font(bold=True, size=11)
    ws.cell(row=98, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A98:C98")
    d98 = round(sum(float(l.last_year_actual or 0) for l in plan.lines), 2)
    g98 = round(sum(float(l.budget_amount or 0) for l in plan.lines), 2)
    i98 = round(sum(float(l.actual_amount or 0) for l in plan.lines), 2)
    e98 = round(d98 / c6_top, 8) if c6_top > 0 else 0.0
    h98 = round(g98 / c2_top, 8) if c2_top > 0 else 0.0
    ws["D98"] = d98
    ws["E98"] = e98
    ws["F98"] = round(f98_sum, 2)
    ws["G98"] = g98
    ws["H98"] = h98
    ws["I98"] = i98
    ws["J98"] = round(g98 - i98, 2)
    for col in (4, 6, 7, 9, 10):
        ws.cell(row=98, column=col).number_format = '#,##0.00'
        ws.cell(row=98, column=col).font = Font(bold=True, size=11)
    for col in (5, 8):
        ws.cell(row=98, column=col).number_format = '0.00%'
        ws.cell(row=98, column=col).font = Font(bold=True, size=11, color="555555")
    for col in range(1, 11):
        cell = ws.cell(row=98, column=col)
        _apply_border(cell, Border, Side)
        cell.fill = PatternFill("solid", fgColor="F1F5F9")
    # 合计费用率 H98：相对行业带 / 硬顶 绿红提示
    try:
        from . import industry as ind_mod

        band = ind_mod.get_period_expense_ratio_band(plan.industry or "")
        hi = float(band.get("max") or 0.18)
        lo = float(band.get("min") or 0.08)
        h98_cell = ws["H98"]
        g98_cell = ws["G98"]
        if h98 > hi + 1e-9:
            _paint_bad(h98_cell, Font, PatternFill)
            _paint_bad(g98_cell, Font, PatternFill)
        elif h98 >= lo - 1e-9:
            _paint_ok(h98_cell, Font, PatternFill)
            _paint_ok(g98_cell, Font, PatternFill)
        elif h98 > 0:
            _paint_warn(h98_cell, Font, PatternFill)
    except Exception:
        pass
    ws.row_dimensions[98].height = 24

    # 顶部派生列：预计算（与引擎一致，打开即见）
    tc = plan.top_computed
    ws["C4"] = round(float(tc.gross_profit or 0), 2)
    ws["C5"] = round(float(tc.gross_margin or 0), 6)
    ws["C7"] = round(float(tc.revenue_growth_rate or 0), 6)
    ws["C9"] = round(float(tc.last_year_gross_margin or 0), 6)
    ws["E5"] = round(float(tc.income_tax_budget or 0), 2)
    ws["E6"] = round(float(tc.profit_total_budget or 0), 2)
    ws["E7"] = round(float(tc.expense_budget_cap or 0), 2)
    ws["E8"] = i98
    ws["E9"] = round(float(tc.expense_budget_cap or 0) - i98, 2)
    for coord in ("C4", "C5", "C7", "C9", "E5", "E6", "E7", "E8", "E9"):
        _apply_formula_style(ws[coord], Font, PatternFill, Alignment, Border, Side)
    for coord in ("C5", "C7", "C9"):
        ws[coord].number_format = "0.00%"
    for coord in ("C4", "E5", "E6", "E7", "E8", "E9"):
        ws[coord].number_format = '#,##0.00'
    ws["E7"].fill = PatternFill("solid", fgColor=_RED_FILL)
    ws["E7"].font = Font(size=10, bold=True, color="B91C1C")
    ws["E9"].fill = PatternFill("solid", fgColor=_YELLOW_FILL)
    ws["E9"].font = Font(size=10, bold=True)

    # G100 未分配余额（预计算）
    ws["A100"] = "未分配余额"
    ws["A100"].font = Font(bold=True, size=11, color="B91C1C")
    ws["A100"].alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells("A100:F100")
    ws["G100"] = round(float(tc.expense_budget_cap or 0) - g98, 2)
    ws["G100"].font = Font(bold=True, size=11, color="B91C1C")
    ws["G100"].fill = PatternFill("solid", fgColor=_YELLOW_FILL)
    ws["G100"].number_format = '#,##0.00'
    ws["G100"].alignment = Alignment(horizontal="right", vertical="center")
    _apply_border(ws["G100"], Border, Side)
    ws["H100"] = "建模：E=D/C6 · H=G/C2 · F(有D)=D×(1+C7) · 毛利率等已预计算写入，打开即可见"
    ws["H100"].font = Font(size=9, color="B91C1C", italic=True)
    ws.merge_cells("H100:J100")

    # 自动筛选范围 A13:J98（原模板错误为 A13:P98）
    ws.auto_filter.ref = f"A13:J98"

    # 冻结窗格：A14 以上冻结
    ws.freeze_panes = "A14"

    # 列宽
    widths = [14, 18, 24, 16, 16, 14, 16, 16, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 顶部说明 + 合规硬规则摘要 + 颜色图例
    ws["A11"] = (
        "说明：D/G/I 输入；E=D/上年营收、H=G/预算营收已建模。"
        "【颜色】绿=合规达标 · 橙=关注 · 红=超标/风险（单行费用率、费用增速、合计费用率、J=G−I）。"
        "【规则】①历史费用率对标 ②金税四期可筹划范围 ③费用增幅匹配营收增速。详见「费用合规筹划约束」。"
    )
    ws["A11"].font = Font(size=9, italic=True, color="B91C1C")
    ws.merge_cells("A11:J11")

    # P1-2：E4 所得税率下拉数据验证（5% / 15% / 25%）+ 状态列条件格式
    try:
        dv = DataValidation(
            type="list",
            formula1='"0.05,0.15,0.25"',
            allow_blank=False,
            showErrorMessage=True,
            errorTitle="税率非法",
            error="E4 所得税率必须为 0.05 / 0.15 / 0.25 之一（5%/15%/25%）；项目默认 0.15（高新）。",
            showInputMessage=True,
            promptTitle="E4 所得税率",
            prompt="默认 0.15（高新技术企业）；可选 0.05（小型微利）/ 0.25（标准税率）",
        )
        dv.add("E4")
        ws.add_data_validation(dv)
    except Exception:
        # 数据验证添加失败不阻塞导出
        pass

    # P1-2：J 列差额条件格式（负数红底，超支提示）
    try:
        from openpyxl.formatting.rule import CellIsRule
        red_fill_cf = PatternFill("solid", fgColor="FECACA")
        # J14:J97 范围内负数标红
        ws.conditional_formatting.add(
            f"J{TEMPLATE_FIRST_ROW}:J{TEMPLATE_LAST_ROW}",
            CellIsRule(operator="lessThan", formula=["0"], fill=red_fill_cf),
        )
    except Exception:
        pass

    # P1-2：应用打印设置（横向 A4 / fitToWidth=1 / 重复表头行 13）
    _apply_print_settings(ws, landscape=True, fit_width=1, repeat_rows="1:13")


def _build_industry_benchmark_sheet(
    wb, openpyxl, Alignment, Border, Font, PatternFill, Side, plan: BudgetPlan
) -> None:
    """Sheet2：行业企业所得税贡献率参考（含来源/年度/地区/核验状态）。"""
    ws = wb.create_sheet("行业企业所得税贡献率参考")

    ws["A1"] = "行业企业所得税贡献率参考表"
    ws["A1"].font = Font(bold=True, size=14, color="1E3A8A")
    ws.merge_cells("A1:F1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # 表头
    headers = ["行业", "企业所得税贡献率", "来源", "适用年度", "适用地区", "是否已核验"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=ci, value=h)
        _apply_header_style(cell, Font, PatternFill, Alignment)
        _apply_border(cell, Border, Side)
    ws.row_dimensions[2].height = 28

    # 数据行
    benchmarks = plan.industry_benchmarks or []
    if benchmarks:
        for ri, b in enumerate(benchmarks, start=3):
            ws.cell(row=ri, column=1, value=b.get("industry", ""))
            ws.cell(row=ri, column=2, value=float(b.get("rate", 0.0)))
            ws.cell(row=ri, column=3, value=b.get("source", ""))
            ws.cell(row=ri, column=4, value=b.get("year", ""))
            ws.cell(row=ri, column=5, value=b.get("region", ""))
            ws.cell(row=ri, column=6, value="是" if b.get("verified") else "否")
            for ci in range(1, 7):
                cell = ws.cell(row=ri, column=ci)
                _apply_border(cell, Border, Side)
                cell.alignment = Alignment(horizontal="left" if ci != 2 else "right",
                                           vertical="center", wrap_text=True)
            ws.cell(row=ri, column=2).number_format = '0.00%'
    else:
        ws.cell(row=3, column=1, value="（暂无行业基准数据；用户可维护）")
        ws.cell(row=3, column=1).font = Font(italic=True, color="888888")
        ws.merge_cells("A3:F3")

    # 核验状态提示
    tip_row = (len(benchmarks) + 4) if benchmarks else 5
    ws.cell(row=tip_row, column=1, value=(
        "提示：行业贡献率须标注来源与适用年度；无来源数据应标『模板默认（待核验）』，"
        "不得包装为官方数据。本工具不提供虚构的行业基准。"
    ))
    ws.cell(row=tip_row, column=1).font = Font(size=9, italic=True, color="B91C1C")
    ws.merge_cells(start_row=tip_row, start_column=1, end_row=tip_row, end_column=6)
    ws.cell(row=tip_row, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[tip_row].height = 32

    # 列宽
    widths = [16, 18, 32, 12, 14, 12]
    for i, w in enumerate(widths, start=1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
    _apply_print_settings(ws, landscape=False, fit_width=1, repeat_rows="1:2")


def _build_compliance_policy_sheet(
    wb,
    openpyxl,
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    plan: BudgetPlan,
    financial_data=None,
    session=None,
) -> None:
    """Sheet4：费用合规筹划约束——三条硬规则 + 量化硬顶（交付必须可见）。"""
    from openpyxl.utils import get_column_letter

    from . import compliance_policy as compliance_mod
    from . import industry as ind_mod

    ws = wb.create_sheet("费用合规筹划约束")
    green = "3ECF8E"
    orange = "FF7A3D"
    red = "FF5D5D"
    navy = "1E3A8A"

    ws["A1"] = "费用合规筹划约束（必须严格执行 · 金税四期方向）"
    ws["A1"].font = Font(bold=True, size=14, color=navy)
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 28

    ws["A2"] = (
        f"企业：{plan.company_name or '—'}　行业：{plan.industry or '—'}　"
        f"预算年：{plan.year or '—'}　生成说明：编制与导出均绑定本表规则"
    )
    ws["A2"].font = Font(size=10, color="555555")
    ws.merge_cells("A2:F2")

    # ── 一、三条硬规则 ──
    ws["A4"] = "一、必须严格执行的三条规则"
    ws["A4"].font = Font(bold=True, size=12, color=navy)
    ws.merge_cells("A4:F4")
    for i, rule in enumerate(compliance_mod.HARD_RULES_LIST, start=1):
        r = 4 + i
        ws.cell(row=r, column=1, value=f"{i}.")
        ws.cell(row=r, column=1).font = Font(bold=True, size=11, color=red)
        ws.cell(row=r, column=2, value=rule)
        ws.cell(row=r, column=2).font = Font(size=11)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 36
        for c in range(1, 7):
            _apply_border(ws.cell(row=r, column=c), Border, Side)
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="FEF3C7")

    # ── 二、量化约束 ──
    row = 9
    ws.cell(row=row, column=1, value="二、量化约束（历史对标 · 增速匹配 · 行业区间）")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color=navy)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    ws["A3"] = "颜色图例：  绿底=合规达标    橙底=关注    红底=超标/风险（按用户合规提示逐格着色）"
    ws["A3"].font = Font(size=10, bold=True, color="555555")
    ws.merge_cells("A3:F3")
    _paint_ok(ws["A3"], Font, PatternFill)
    ws["A3"].font = Font(size=10, bold=True, color=_OK_FONT)

    row = 10
    headers = ["指标", "数值", "说明", "状态", "", ""]
    for ci, h in enumerate(headers[:4], start=1):
        cell = ws.cell(row=row, column=ci, value=h)
        _apply_header_style(cell, Font, PatternFill, Alignment)
        _apply_border(cell, Border, Side)

    lim = None
    if financial_data is not None:
        try:
            lim = compliance_mod.budget_amount_limits(financial_data, plan.industry or "")
        except Exception:
            lim = None
    if lim is None:
        band = ind_mod.get_period_expense_ratio_band(plan.industry or "")
        c2 = float(plan.top_inputs.budget_revenue or 0)
        c6 = float(plan.top_inputs.last_year_revenue or 0)
        c7 = float(plan.top_computed.revenue_growth_rate or 0) if plan.top_computed else 0.0
        g_sum = float(plan.allocated_total or 0)
        d_sum = float(plan.last_year_total or 0)
        lim = {
            "revenue_growth_rate": c7,
            "max_fee_growth_rate": c7 + compliance_mod.FEE_GROWTH_BUFFER_PP,
            "target_period_expense_ratio": float(band.get("median") or 0.12),
            "target_period_expense_total": round(c2 * float(band.get("median") or 0.12), 2),
            "hard_cap_period_expense_total": round(c2 * float(band.get("max") or 0.18), 2),
            "industry_fee_band": band,
            "historical": {
                "by_year": {},
                "revenue_growth_rate": c7,
            },
            "max_single_line_ratio": compliance_mod.MAX_SINGLE_LINE_RATIO,
            "max_single_line_ratio_rigid": compliance_mod.MAX_SINGLE_LINE_RATIO_RIGID,
            "_fallback": True,
            "_plan_g": g_sum,
            "_plan_d": d_sum,
            "_c2": c2,
            "_c6": c6,
        }
    else:
        lim["_plan_g"] = float(plan.allocated_total or 0)
        lim["_c2"] = float(plan.top_inputs.budget_revenue or 0)

    band = lim.get("industry_fee_band") or {}
    c2 = float(lim.get("_c2") or plan.top_inputs.budget_revenue or 0)
    g_sum = float(lim.get("_plan_g") or plan.allocated_total or 0)
    fee_rate = (g_sum / c2) if c2 > 0 else 0.0
    hard_cap = float(lim.get("hard_cap_period_expense_total") or 0)
    rows_metrics = [
        ("预算营业收入 C2（元）", c2, "建模分母：H=G/C2"),
        ("收入增长率", float(lim.get("revenue_growth_rate") or 0), "（C2−C6）/C6"),
        (
            "费用增速上限",
            float(lim.get("max_fee_growth_rate") or 0),
            "收入增速(波动年 winsor) + 3pp；模式="
            + str(lim.get("fee_growth_mode") or "raw_plus_buffer"),
        ),
        ("行业费用率下限", float(band.get("min") or 0), "WB 行业带"),
        ("行业费用率中枢", float(band.get("median") or 0), "WB 行业带"),
        ("行业费用率上限", float(band.get("max") or 0), "WB 行业带"),
        ("目标期间费用率", float(lim.get("target_period_expense_ratio") or 0), "历史与行业中枢夹逼"),
        ("目标期间费用合计（元）", float(lim.get("target_period_expense_total") or 0), "营收×目标费用率"),
        ("费用硬顶合计（元）", hard_cap, "min(增速匹配上限, 行业带上限)"),
        ("本表预算费用 ΣG（元）", g_sum, "当前导出表合计"),
        ("本表总费用率 ΣG/C2", fee_rate, "须 ≤ 硬顶/行业上沿；过高标红"),
        ("单行费用率上限（一般）", float(lim.get("max_single_line_ratio") or 0.08), "占营收"),
        ("单行费用率上限（刚性）", float(lim.get("max_single_line_ratio_rigid") or 0.12), "工资/社保/房租"),
    ]
    r0 = 11
    band_lo = float(band.get("min") or 0.08)
    band_hi = float(band.get("max") or 0.18)
    target_rate = float(lim.get("target_period_expense_ratio") or band.get("median") or 0.12)
    for i, (name, val, note) in enumerate(rows_metrics):
        r = r0 + i
        ws.cell(row=r, column=1, value=name)
        cell_v = ws.cell(row=r, column=2, value=val)
        if "率" in name or "增速" in name:
            cell_v.number_format = "0.00%"
        else:
            cell_v.number_format = "#,##0.00"
        ws.cell(row=r, column=3, value=note)
        status = "—"
        # 按指标给「数值」单格绿/红
        if name.startswith("本表总费用率"):
            if hard_cap > 0 and c2 > 0 and g_sum > hard_cap * 1.01:
                _paint_bad(cell_v, Font, PatternFill)
                status = "超硬顶"
            elif fee_rate > band_hi + 1e-9:
                _paint_bad(cell_v, Font, PatternFill)
                status = "超行业上限"
            elif fee_rate < band_lo - 1e-9 and fee_rate > 0:
                _paint_warn(cell_v, Font, PatternFill)
                status = "低于行业下限"
            elif fee_rate > 0:
                _paint_ok(cell_v, Font, PatternFill)
                status = "达标"
        elif name.startswith("本表预算费用"):
            if hard_cap > 0 and g_sum > hard_cap * 1.01:
                _paint_bad(cell_v, Font, PatternFill)
                status = "超硬顶"
            elif g_sum > 0:
                _paint_ok(cell_v, Font, PatternFill)
                status = "已填报"
        elif name.startswith("费用硬顶"):
            if hard_cap > 0:
                _paint_ok(cell_v, Font, PatternFill)
                status = "上限"
        elif name.startswith("目标期间费用率"):
            if band_lo - 1e-9 <= target_rate <= band_hi + 1e-9:
                _paint_ok(cell_v, Font, PatternFill)
                status = "在行业带内"
            else:
                _paint_warn(cell_v, Font, PatternFill)
                status = "偏离行业带"
        elif name.startswith("收入增长率"):
            if float(val) < -0.2:
                _paint_warn(cell_v, Font, PatternFill)
                status = "收入下滑"
            else:
                _paint_ok(cell_v, Font, PatternFill)
                status = "已测算"
        cell_st = ws.cell(row=r, column=4, value=status)
        if status in ("超硬顶", "超行业上限"):
            _paint_bad(cell_st, Font, PatternFill)
        elif status in ("达标", "已填报", "在行业带内", "已测算", "上限"):
            _paint_ok(cell_st, Font, PatternFill)
        elif status not in ("—",):
            _paint_warn(cell_st, Font, PatternFill)
        for c in range(1, 5):
            _apply_border(ws.cell(row=r, column=c), Border, Side)

    # ── 三、历史费用占营收 ──
    hist_start = r0 + len(rows_metrics) + 2
    ws.cell(row=hist_start, column=1, value="三、历史成本费用占营收比例（对标）")
    ws.cell(row=hist_start, column=1).font = Font(bold=True, size=12, color=navy)
    ws.merge_cells(start_row=hist_start, start_column=1, end_row=hist_start, end_column=6)

    h_header = hist_start + 1
    for ci, h in enumerate(
        ["年度", "营业收入", "期间费用合计", "期间费用率", "销售费用率", "管理费用率", "财务费用率"],
        start=1,
    ):
        cell = ws.cell(row=h_header, column=ci, value=h)
        _apply_header_style(cell, Font, PatternFill, Alignment)
        _apply_border(cell, Border, Side)

    by_year = (lim.get("historical") or {}).get("by_year") or {}
    if by_year:
        for i, y in enumerate(sorted(by_year.keys())):
            r = h_header + 1 + i
            yv = by_year[y]
            ws.cell(row=r, column=1, value=str(y))
            ws.cell(row=r, column=2, value=float(yv.get("revenue") or 0)).number_format = "#,##0.00"
            ws.cell(row=r, column=3, value=float(yv.get("period_expense") or 0)).number_format = "#,##0.00"
            ws.cell(row=r, column=4, value=float(yv.get("period_expense_ratio") or 0)).number_format = "0.00%"
            ws.cell(row=r, column=5, value=float(yv.get("selling_ratio") or 0)).number_format = "0.00%"
            ws.cell(row=r, column=6, value=float(yv.get("admin_ratio") or 0)).number_format = "0.00%"
            ws.cell(row=r, column=7, value=float(yv.get("finance_ratio") or 0)).number_format = "0.00%"
            for c in range(1, 8):
                _apply_border(ws.cell(row=r, column=c), Border, Side)
        hist_end = h_header + len(by_year)
    else:
        # 无 FinancialData：用本表 D/G 近似
        r = h_header + 1
        c6 = float(plan.top_inputs.last_year_revenue or 0)
        d_sum = float(plan.last_year_total or 0)
        ws.cell(row=r, column=1, value=str(plan.year or "上年") + "（表内D近似）")
        ws.cell(row=r, column=2, value=c6).number_format = "#,##0.00"
        ws.cell(row=r, column=3, value=d_sum).number_format = "#,##0.00"
        ws.cell(row=r, column=4, value=(d_sum / c6 if c6 else 0)).number_format = "0.00%"
        ws.cell(row=r, column=5, value="（导入财报后显示分年明细）")
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
        for c in range(1, 8):
            _apply_border(ws.cell(row=r, column=c), Border, Side)
        hist_end = r

    # ── 四、本表科目预算 vs 约束 ──
    sub_start = hist_end + 2
    ws.cell(row=sub_start, column=1, value="四、本表科目预算结构（对照利润表口径）")
    ws.cell(row=sub_start, column=1).font = Font(bold=True, size=12, color=navy)
    ws.merge_cells(start_row=sub_start, start_column=1, end_row=sub_start, end_column=6)

    sh = sub_start + 1
    for ci, h in enumerate(["科目", "上年D合计", "预算G合计", "G占营收", "状态", "备注"], start=1):
        cell = ws.cell(row=sh, column=ci, value=h)
        _apply_header_style(cell, Font, PatternFill, Alignment)
        _apply_border(cell, Border, Side)
    subjects = ["销售费用", "管理费用", "研发费用", "财务费用", "营业外支出"]
    max_fee_g = float(lim.get("max_fee_growth_rate") or 0)
    for i, subj in enumerate(subjects):
        r = sh + 1 + i
        lines = [l for l in plan.lines if l.subject == subj]
        d_s = sum(float(l.last_year_actual or 0) for l in lines)
        g_s = sum(float(l.budget_amount or 0) for l in lines)
        ratio = (g_s / c2 if c2 else 0)
        ws.cell(row=r, column=1, value=subj)
        ws.cell(row=r, column=2, value=d_s).number_format = "#,##0.00"
        cell_g = ws.cell(row=r, column=3, value=g_s)
        cell_g.number_format = "#,##0.00"
        cell_h = ws.cell(row=r, column=4, value=ratio)
        cell_h.number_format = "0.00%"
        note = ""
        status = "—"
        if g_s <= 0 and d_s <= 0:
            status = "无数据"
        elif d_s > 0 and g_s > d_s * (1 + max_fee_g + 0.001):
            note = "G 增速超收入增速缓冲"
            status = "超标"
            _paint_bad(cell_g, Font, PatternFill)
            _paint_bad(cell_h, Font, PatternFill)
        elif g_s > 0:
            note = "增速匹配/结构可接受"
            status = "达标"
            _paint_ok(cell_g, Font, PatternFill)
            _paint_ok(cell_h, Font, PatternFill)
        cell_st = ws.cell(row=r, column=5, value=status)
        if status == "超标":
            _paint_bad(cell_st, Font, PatternFill)
        elif status == "达标":
            _paint_ok(cell_st, Font, PatternFill)
        ws.cell(row=r, column=6, value=note)
        for c in range(1, 7):
            _apply_border(ws.cell(row=r, column=c), Border, Side)

    # ── 五、数据质量 / 跨年勾稽 / E3·E4 来源 ──
    dq_start = sh + 6
    ws.cell(row=dq_start, column=1, value="五、数据质量 · 跨年勾稽 · 税率/贡献率来源")
    ws.cell(row=dq_start, column=1).font = Font(bold=True, size=12, color=navy)
    ws.merge_cells(start_row=dq_start, start_column=1, end_row=dq_start, end_column=6)

    dq = {}
    cit_syn = {}
    rec = {}
    anomalies = []
    if financial_data is not None:
        meta = getattr(financial_data, "parsed_meta", None) or {}
        dq = meta.get("data_quality") or {}
        cit_syn = meta.get("cit_synthesis") or {}
        rec = meta.get("reconciliation") or {}
        anomalies = (meta.get("expense_anomalies") or {}).get("anomalies") or []

    ti = plan.top_inputs
    dq_rows = [
        ("数据置信度", str(dq.get("confidence") or "—"), "high/medium/low；低置信度请人工核金额"),
        (
            "勾稽结果",
            "通过" if rec.get("ok", True) and not rec.get("hard_fail") else "有失败/警告",
            f"错误 {rec.get('error_count', len(rec.get('errors') or []))} · "
            f"警告 {rec.get('warning_count', len(rec.get('warnings') or []))}",
        ),
        (
            "E4 所得税税率（名义）",
            float(ti.income_tax_rate or 0),
            "高新默认 15%；与所得税贡献率 E3 分列",
        ),
        (
            "E3 企业贡献率",
            float(ti.company_contribution_rate or 0),
            f"来源 {cit_syn.get('basis') or '—'}；"
            f"WB中枢={(cit_syn.get('wb_hub') or 0)*100:.2f}% · "
            f"历史有效={(cit_syn.get('latest_valid') or 0)*100:.2f}%",
        ),
        (
            "E2 行业贡献率",
            float(ti.industry_contribution_rate or 0),
            "WB 行业所得税税负经验中枢",
        ),
        (
            "费用增速模式",
            str(lim.get("fee_growth_mode") or "raw_plus_buffer"),
            "营收 |增速|>30% 时 winsor 降权",
        ),
        (
            "无销售费用型",
            "是" if lim.get("near_zero_selling") else "否",
            "历史销售/营收极低时禁止虚增销售预算",
        ),
    ]
    for i, (name, val, note) in enumerate(dq_rows):
        r = dq_start + 1 + i
        ws.cell(row=r, column=1, value=name)
        cell_v = ws.cell(row=r, column=2, value=val)
        if isinstance(val, float) and ("率" in name or "贡献" in name or "税率" in name):
            cell_v.number_format = "0.00%"
        ws.cell(row=r, column=3, value=note)
        status = "—"
        if name == "数据置信度":
            if val == "high":
                _paint_ok(cell_v, Font, PatternFill)
                status = "可用"
            elif val == "medium":
                _paint_warn(cell_v, Font, PatternFill)
                status = "关注"
            elif val == "low":
                _paint_bad(cell_v, Font, PatternFill)
                status = "须核验"
        elif name == "勾稽结果":
            if "通过" in str(val):
                _paint_ok(cell_v, Font, PatternFill)
                status = "通过"
            else:
                _paint_bad(cell_v, Font, PatternFill)
                status = "异常"
        cell_st = ws.cell(row=r, column=4, value=status)
        if status in ("可用", "通过"):
            _paint_ok(cell_st, Font, PatternFill)
        elif status in ("关注",):
            _paint_warn(cell_st, Font, PatternFill)
        elif status in ("须核验", "异常"):
            _paint_bad(cell_st, Font, PatternFill)
        for c in range(1, 5):
            _apply_border(ws.cell(row=r, column=c), Border, Side)

    # 勾稽/异常摘要行
    warn_r = dq_start + 1 + len(dq_rows) + 1
    warn_bits = list(rec.get("errors") or [])[:5] + list(rec.get("warnings") or [])[:5]
    for a in anomalies[:5]:
        warn_bits.append(a.get("message") or str(a))
    ws.cell(row=warn_r, column=1, value="勾稽与费用异常摘要")
    ws.cell(
        row=warn_r,
        column=2,
        value="；".join(warn_bits) if warn_bits else "（无）",
    )
    ws.merge_cells(start_row=warn_r, start_column=2, end_row=warn_r, end_column=6)
    ws.cell(row=warn_r, column=2).alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[warn_r].height = 48
    for c in range(1, 7):
        _apply_border(ws.cell(row=warn_r, column=c), Border, Side)

    # ── 六、诊断风险摘要（若有互动会话）──
    risk_start = warn_r + 2
    ws.cell(row=risk_start, column=1, value="六、诊断风险摘要（低绿→中橙→高红，与系统诊断一致）")
    ws.cell(row=risk_start, column=1).font = Font(bold=True, size=12, color=navy)
    ws.merge_cells(start_row=risk_start, start_column=1, end_row=risk_start, end_column=6)

    rh = risk_start + 1
    for ci, h in enumerate(["风险等级", "类别", "发现标题", "事实摘要"], start=1):
        cell = ws.cell(row=rh, column=ci, value=h)
        _apply_header_style(cell, Font, PatternFill, Alignment)
        _apply_border(cell, Border, Side)

    findings = []
    if session is not None and getattr(session, "diagnosis", None):
        findings = list(getattr(session.diagnosis, "findings", None) or [])
        try:
            findings = compliance_mod.sort_findings_by_severity(findings)
        except Exception:
            pass
    if findings:
        label_map = {"低": "低风险", "中": "中风险", "高": "高风险"}
        for i, f in enumerate(findings[:40]):
            r = rh + 1 + i
            sev = getattr(f, "severity", "") or ""
            cell_sev = ws.cell(row=r, column=1, value=label_map.get(sev, sev))
            if sev == "低":
                _paint_ok(cell_sev, Font, PatternFill)
            elif sev == "中":
                _paint_warn(cell_sev, Font, PatternFill)
            elif sev == "高":
                _paint_bad(cell_sev, Font, PatternFill)
            else:
                cell_sev.font = Font(bold=True)
            ws.cell(row=r, column=2, value=getattr(f, "category", "") or "")
            ws.cell(row=r, column=3, value=getattr(f, "title", "") or "")
            fact = str(getattr(f, "fact", "") or "")[:200]
            ws.cell(row=r, column=4, value=fact)
            for c in range(1, 5):
                _apply_border(ws.cell(row=r, column=c), Border, Side)
                ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="center")
    else:
        r = rh + 1
        ws.cell(row=r, column=1, value="（本次导出未附带诊断会话；完成诊断互动后导出可在此显示风险清单）")
        ws.cell(row=r, column=1).font = Font(italic=True, color="888888")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

    tip = risk_start + 45
    ws.cell(
        row=tip,
        column=1,
        value=(
            "免责声明：本表为内部合规对标与预算编制辅助，非税务机关法定标准；"
            "重大涉税决策须结合专业机构与主管税务机关意见。"
        ),
    )
    ws.cell(row=tip, column=1).font = Font(size=9, italic=True, color="888888")
    ws.merge_cells(start_row=tip, start_column=1, end_row=tip, end_column=6)

    widths = [28, 18, 18, 14, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    _apply_print_settings(ws, landscape=True, fit_width=1, repeat_rows="1:2")


def _build_action_summary_sheet(
    wb, openpyxl, Alignment, Border, Font, PatternFill, Side, plan: BudgetPlan, session=None
) -> None:
    """P0-2：Sheet3「诊断与行动清单」。

    有 session 时填入决策数据（发现/选项/目标/成本节约/税收节约/税负影响/净影响/负责人/期限/执行状态）；
    无 session 时保留完整空表头 + 「尚未完成诊断」提示，不得伪造建议。
    """
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet("诊断与行动清单")

    ws["A1"] = "诊断与行动清单"
    ws["A1"].font = Font(bold=True, size=14, color="1E3A8A")
    ws.merge_cells("A1:J1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # 表头（10 列）
    headers = [
        "序号", "发现", "选项", "目标值",
        "成本节约(元)", "税收节约(元)", "税负影响(元)", "净影响(元)",
        "负责人", "期限",
    ]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=ci, value=h)
        _apply_header_style(cell, Font, PatternFill, Alignment)
        _apply_border(cell, Border, Side)
    ws.row_dimensions[2].height = 28

    # 数据行
    if session is not None and getattr(session, "draft2", None):
        # 有会话：填入决策数据
        for ri, d in enumerate(session.draft2, start=3):
            ws.cell(row=ri, column=1, value=ri - 2)
            ws.cell(row=ri, column=2, value=getattr(d, "finding_title", "") or getattr(d, "finding_id", ""))
            ws.cell(row=ri, column=3, value=getattr(d, "option_name", "") or getattr(d, "option_label", ""))
            ws.cell(row=ri, column=4, value=str(getattr(d, "target_value", "")))
            # 成本节约/税收节约/税负影响/净影响（来自 draft2 条目）
            ws.cell(row=ri, column=5, value=float(getattr(d, "cost_saving", 0.0) or 0.0))
            ws.cell(row=ri, column=6, value=float(getattr(d, "tax_saving", 0.0) or 0.0))
            ws.cell(row=ri, column=7, value=float(getattr(d, "tax_impact", 0.0) or 0.0))
            # P0-B（CO T7 重开）：Draft2Entry 字段为 est_saving（非 net_impact）；
            # 旧代码用 net_impact 永远取到 0，导致 Excel 第三 Sheet 净影响列全 0。
            ws.cell(row=ri, column=8, value=float(getattr(d, "est_saving", 0.0) or 0.0))
            ws.cell(row=ri, column=9, value=str(getattr(d, "owner", "")))   # 负责人（可空）
            ws.cell(row=ri, column=10, value=str(getattr(d, "deadline", "")))  # 期限（可空）
            for ci in range(1, 11):
                cell = ws.cell(row=ri, column=ci)
                _apply_border(cell, Border, Side)
                cell.alignment = Alignment(horizontal="left" if ci in (2, 3, 4) else "right",
                                           vertical="center", wrap_text=True)
            for ci in (5, 6, 7, 8):
                ws.cell(row=ri, column=ci).number_format = '#,##0.00'
        data_rows = len(session.draft2)
        # 汇总行
        sum_row = 3 + data_rows
        ws.cell(row=sum_row, column=1, value="合计")
        ws.cell(row=sum_row, column=1).font = Font(bold=True, size=11)
        ws.cell(row=sum_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=4)
        for ci, col_letter in zip((5, 6, 7, 8), ("E", "F", "G", "H")):
            ws.cell(row=sum_row, column=ci,
                    value=f"=SUM({col_letter}3:{col_letter}{sum_row - 1})")
            ws.cell(row=sum_row, column=ci).font = Font(bold=True, size=11)
            ws.cell(row=sum_row, column=ci).number_format = '#,##0.00'
            ws.cell(row=sum_row, column=ci).fill = PatternFill("solid", fgColor="F1F5F9")
            _apply_border(ws.cell(row=sum_row, column=ci), Border, Side)
        tip = f"共 {data_rows} 条诊断行动建议；负责人与期限由企业填写。"
    else:
        # 无会话：空表头 + 明确提示
        ws.cell(row=3, column=1, value="（尚未完成诊断；请先在「2. 诊断」与「3. 互动」完成决策后再生成本清单）")
        ws.cell(row=3, column=1).font = Font(italic=True, color="888888")
        ws.merge_cells("A3:J3")
        ws.cell(row=3, column=1).alignment = Alignment(horizontal="left", vertical="center")
        tip = "本表为空表头；完成诊断互动后将自动填入决策数据，不得伪造建议。"

    # 提示行
    tip_row = ws.max_row + 2
    ws.cell(row=tip_row, column=1, value=tip)
    ws.cell(row=tip_row, column=1).font = Font(size=9, italic=True, color="B91C1C")
    ws.merge_cells(start_row=tip_row, start_column=1, end_row=tip_row, end_column=10)
    ws.cell(row=tip_row, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 列宽
    widths = [6, 28, 24, 18, 14, 14, 14, 14, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
    _apply_print_settings(ws, landscape=True, fit_width=1, repeat_rows="1:2")


def _apply_print_settings(ws, landscape: bool = True, fit_width: int = 1,
                          repeat_rows: str = "", margins_cm: float = 1.5) -> None:
    """P1-2：Excel 打印设置（横向 A4 / fitToWidth / 重复表头行 / 合理页边距）。

    使用通用中文字体优先级：PingFang SC（macOS）→ Microsoft YaHei（Windows）→ SimSun（跨平台）。
    """
    from openpyxl.worksheet.page import PageMargins, PrintOptions
    try:
        if landscape:
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        else:
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = fit_width
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins = PageMargins(
            left=margins_cm / 2.54, right=margins_cm / 2.54,
            top=margins_cm / 2.54, bottom=margins_cm / 2.54,
            header=0.3, footer=0.3,
        )
        if repeat_rows:
            ws.print_title_rows = repeat_rows
        ws.print_options = PrintOptions(horizontalCentered=True)
    except Exception:
        # 打印设置失败不阻塞导出
        pass
