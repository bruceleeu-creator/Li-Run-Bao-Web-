"""利润宝 · 三 Sheet 预算模板导出填充（DeepSeek 可选增强）。

产出与 `budget_3sheet_WB-CO-TR-20260726.xlsx` 同构的工作簿：
1. 费用预算表（A1:J100，公式列 E/F/H/J 保留）
2. 行业企业所得税贡献率参考
3. 诊断与行动清单

流程：
- 规则：从 FinancialData 提取营收/成本，科目余额表/费用科目映射到 84 行
- AI（可选）：DeepSeek 识别顶部 4 指标 + 将费用金额分配到模板行
- 写出：core.budget_template.write_template（禁止覆盖原模板）
"""
from __future__ import annotations

import json
import logging
import os
import re
from pathlib import Path
from typing import Any, Optional, Sequence

from . import budget as budget_mod
from . import budget_template as tpl_mod
from .models import FinancialData

logger = logging.getLogger(__name__)

# 项目内标准样例模板（与用户提供的格式一致）
PROJECT_ROOT = Path(__file__).resolve().parent.parent
CANONICAL_TEMPLATE = (
    PROJECT_ROOT / "demo_output" / "budget_3sheet_WB-CO-TR-20260726.xlsx"
)

# 科目余额表 → 模板费用名称关键词（确定性映射，可复核）
_LEDGER_TO_EXPENSE_KEYWORDS: list[tuple[str, list[str]]] = [
    ("业务招待费", ["业务招待费"]),
    ("广告宣传费", ["广宣费", "广告"]),
    ("咨询服务费", ["咨询", "中介机构服务费"]),
    ("福利费", ["职工福利费"]),
    ("教育经费", ["职工教育经费"]),
    ("研发费用", ["研发费用"]),
    ("折旧", ["摊销折旧费", "折旧费"]),
]

_INCOME_TO_SUBJECT: list[tuple[str, str]] = [
    ("销售费用", "销售费用"),
    ("管理费用", "管理费用"),
    ("财务费用", "财务费用"),
    ("研发费用", "管理费用"),  # 模板中研发在管理费用大类下
]


def _amount(data: FinancialData, statement: str, account: str, year: int) -> float:
    if statement == "income":
        table = data.income_statement
    elif statement == "balance":
        table = data.balance_sheet
    else:
        table = data.account_balances
    val = (table.get(account) or {}).get(year)
    try:
        return float(val or 0.0)
    except (TypeError, ValueError):
        return 0.0




def extract_period_expenses(data: FinancialData) -> dict:
    """从利润表提取最新年/上年期间费用，供 DeepSeek 占比对账。"""
    years = sorted(data.years or [])
    latest = years[-1] if years else 0
    prev = years[-2] if len(years) >= 2 else latest
    return {
        "selling_latest": _amount(data, "income", "销售费用", latest),
        "admin_latest": _amount(data, "income", "管理费用", latest),
        "rd_latest": _amount(data, "income", "研发费用", latest),
        "finance_latest": _amount(data, "income", "财务费用", latest),
        "selling_prev": _amount(data, "income", "销售费用", prev),
        "admin_prev": _amount(data, "income", "管理费用", prev),
        "rd_prev": _amount(data, "income", "研发费用", prev),
        "finance_prev": _amount(data, "income", "财务费用", prev),
        "year_latest": latest,
        "year_prev": prev,
    }

def _fill_top_from_data(plan: budget_mod.BudgetPlan, data: FinancialData) -> str:
    from . import pipeline as pipeline_mod

    years = sorted(data.years or [])
    latest = years[-1] if years else 0
    prev = years[-2] if len(years) >= 2 else latest
    ti = plan.top_inputs
    ti.budget_revenue = _amount(data, "income", "营业收入", latest)
    ti.budget_cost = _amount(data, "income", "营业成本", latest)
    ti.last_year_revenue = _amount(data, "income", "营业收入", prev)
    ti.last_year_cost = _amount(data, "income", "营业成本", prev)
    plan.year = latest
    plan.company_name = data.company_name or plan.company_name
    # Policy 单点：E2/E3/E4 + 行业（会话导入时已有；否则现算）
    policy = pipeline_mod.policy_from_data(data)
    pol_notes = pipeline_mod.apply_policy_to_plan(plan, policy)
    note = f"结构化提取 · 预算年 {latest}，上年 {prev} · 行业={plan.industry}"
    if pol_notes:
        note += " · " + "；".join(pol_notes)
    return note


def apply_top_indicators(plan: budget_mod.BudgetPlan, indicators: dict) -> None:
    """把识别结果写入顶部 TopInputs（供 Web 层 AI 回调使用）。"""
    ti = plan.top_inputs
    ti.budget_revenue = float(indicators.get("budget_revenue", 0) or 0)
    ti.budget_cost = float(indicators.get("budget_cost", 0) or 0)
    ti.last_year_revenue = float(indicators.get("last_year_revenue", 0) or 0)
    ti.last_year_cost = float(indicators.get("last_year_cost", 0) or 0)


def _match_lines(plan: budget_mod.BudgetPlan, keywords: list[str]) -> list[budget_mod.ExpenseLine]:
    hits = []
    for line in plan.lines:
        blob = f"{line.expense_name}{line.invoice_name}{line.subject}"
        if any(k in blob for k in keywords):
            hits.append(line)
    return hits


def _distribute(amount: float, lines: list[budget_mod.ExpenseLine], field: str) -> None:
    """把金额均分到匹配行（确定性，可复核）。"""
    if amount <= 0 or not lines:
        return
    each = round(amount / len(lines), 2)
    # 尾差放第一行
    for i, line in enumerate(lines):
        val = each if i < len(lines) - 1 else round(amount - each * (len(lines) - 1), 2)
        setattr(line, field, float(getattr(line, field, 0.0) or 0.0) + val)


def _fill_lines_from_data(plan: budget_mod.BudgetPlan, data: FinancialData) -> str:
    """用利润表期间费用 + 科目余额表确定性灌入 D/G/I。"""
    years = sorted(data.years or [])
    latest = years[-1] if years else 0
    prev = years[-2] if len(years) >= 2 else latest
    notes = []

    # 1) 科目余额表精确项
    for ledger_name, keywords in _LEDGER_TO_EXPENSE_KEYWORDS:
        prev_amt = _amount(data, "ledger", ledger_name, prev)
        latest_amt = _amount(data, "ledger", ledger_name, latest)
        lines = _match_lines(plan, keywords)
        if not lines:
            continue
        if prev_amt > 0:
            _distribute(prev_amt, lines, "last_year_actual")
            _distribute(prev_amt, lines, "budget_amount")  # 默认预算=上年
        if latest_amt > 0:
            _distribute(latest_amt, lines, "actual_amount")
            if prev_amt <= 0:
                _distribute(latest_amt, lines, "budget_amount")
        if prev_amt or latest_amt:
            notes.append(f"{ledger_name}→{len(lines)}行")

    # 2) 利润表期间费用：保证大类合计被灌入（余额表不足时补差额到科目代表行）
    for income_acc, subject in _INCOME_TO_SUBJECT:
        prev_amt = _amount(data, "income", income_acc, prev)
        latest_amt = _amount(data, "income", income_acc, latest)
        if income_acc == "研发费用":
            subject_lines = _match_lines(plan, ["研发费用"]) or [
                l for l in plan.lines if l.subject == subject
            ][:1]
        else:
            subject_lines = [l for l in plan.lines if l.subject == subject]
        if not subject_lines:
            continue
        already_prev = sum(float(l.last_year_actual or 0) for l in subject_lines)
        already_act = sum(float(l.actual_amount or 0) for l in subject_lines)
        # 代表行集：按关键词权重取前 N 行，避免整笔只落一行导致明细全空
        def _w(line: budget_mod.ExpenseLine) -> int:
            blob = f"{line.expense_name}{line.invoice_name}"
            score = 1
            for kws, wt in (
                (["工资", "薪酬", "奖金", "社保", "公积金"], 12),
                (["广宣", "推广", "广告"], 9),
                (["招待"], 7),
                (["差旅"], 6),
                (["房租", "租赁"], 10),
                (["利息", "手续"], 8),
                (["研发"], 8),
                (["折旧", "摊销"], 6),
            ):
                if any(k in blob for k in kws):
                    score = max(score, wt)
            return score

        ranked = sorted(subject_lines, key=_w, reverse=True)
        targets = ranked[: min(8, len(ranked))]
        weights = [_w(l) for l in targets]
        wsum = sum(weights) or 1

        # 补足上年：差额按权重拆到多行
        if prev_amt > already_prev + 1:
            gap = prev_amt - already_prev
            for l, w in zip(targets, weights):
                add = round(gap * (w / wsum), 2)
                l.last_year_actual = float(l.last_year_actual or 0) + add
                if float(l.budget_amount or 0) <= 0:
                    l.budget_amount = float(l.last_year_actual or 0)
                if float(l.reference_amount or 0) <= 0:
                    l.reference_amount = float(l.budget_amount or 0)
            notes.append(f"{income_acc}上年补{gap:,.0f}→{len(targets)}行")
        # 补足最新年实际
        if latest_amt > already_act + 1:
            gap = latest_amt - already_act
            for l, w in zip(targets, weights):
                add = round(gap * (w / wsum), 2)
                l.actual_amount = float(l.actual_amount or 0) + add
                if float(l.budget_amount or 0) <= 0:
                    l.budget_amount = float(l.actual_amount or 0)
                if float(l.reference_amount or 0) <= 0 and float(l.last_year_actual or 0) <= 0:
                    l.reference_amount = float(l.budget_amount or 0)
            notes.append(f"{income_acc}本年补{gap:,.0f}→{len(targets)}行")
        # 完全空白：整笔拆到多行
        already_prev2 = sum(float(l.last_year_actual or 0) for l in subject_lines)
        already_act2 = sum(float(l.actual_amount or 0) for l in subject_lines)
        if prev_amt > 0 and already_prev2 <= 0:
            for l, w in zip(targets, weights):
                l.last_year_actual = round(prev_amt * (w / wsum), 2)
                l.budget_amount = float(l.last_year_actual)
                l.reference_amount = float(l.budget_amount)
            notes.append(f"{income_acc}上年拆分→{len(targets)}行")
        if latest_amt > 0 and already_act2 <= 0:
            for l, w in zip(targets, weights):
                l.actual_amount = round(latest_amt * (w / wsum), 2)
                if float(l.budget_amount or 0) <= 0:
                    l.budget_amount = float(l.actual_amount)
                if float(l.reference_amount or 0) <= 0:
                    l.reference_amount = float(l.budget_amount)
            notes.append(f"{income_acc}本年拆分→{len(targets)}行")

    return "规则映射 · " + ("；".join(notes) if notes else "无费用明细可映射")


def apply_period_totals_to_plan(plan: budget_mod.BudgetPlan, period: dict) -> str:
    """用 AI 提取的期间费用合计，补齐代表行（当明细仍空时）。"""
    if not period:
        return ""
    mapping = [
        ("selling_expense", "prev_selling_expense", "销售费用", ["销售费用"]),
        ("admin_expense", "prev_admin_expense", "管理费用", ["管理费用"]),
        ("rd_expense", "prev_rd_expense", "管理费用", ["研发费用"]),
        ("finance_expense", "prev_finance_expense", "财务费用", ["财务费用"]),
    ]
    notes = []
    for act_key, prev_key, subject, keywords in mapping:
        try:
            act = float(period.get(act_key) or 0)
            prev = float(period.get(prev_key) or 0)
        except (TypeError, ValueError):
            act, prev = 0.0, 0.0
        lines = _match_lines(plan, keywords) or [l for l in plan.lines if l.subject == subject]
        if not lines:
            continue
        primary = lines[0]
        if prev > 0 and sum(float(l.last_year_actual or 0) for l in lines) <= 0:
            primary.last_year_actual = prev
            primary.budget_amount = prev
            notes.append(f"AI期间{prev_key}→R{primary.row}")
        if act > 0 and sum(float(l.actual_amount or 0) for l in lines) <= 0:
            primary.actual_amount = act
            if float(primary.budget_amount or 0) <= 0:
                primary.budget_amount = act
            notes.append(f"AI期间{act_key}→R{primary.row}")
    return "；".join(notes)


def apply_line_allocations(plan: budget_mod.BudgetPlan, items: Sequence[dict]) -> int:
    """应用 AI/外部返回的行分配 JSON 列表，返回成功写入行数。"""
    by_row = {l.row: l for l in plan.lines}
    applied = 0
    for item in items:
        if not isinstance(item, dict):
            continue
        try:
            row = int(item.get("row"))
        except (TypeError, ValueError):
            continue
        line = by_row.get(row)
        if line is None:
            continue
        for field, key in (
            ("last_year_actual", "last_year_actual"),
            ("budget_amount", "budget_amount"),
            ("actual_amount", "actual_amount"),
        ):
            raw = item.get(key)
            if raw is None:
                continue
            try:
                val = float(raw)
            except (TypeError, ValueError):
                m = re.search(r"[-+]?\d+(?:\.\d+)?", str(raw).replace(",", ""))
                val = float(m.group(0)) if m else None
            if val is None or val < 0:
                continue
            setattr(line, field, val)
        applied += 1
    return applied


def _subject_weight(line: budget_mod.ExpenseLine) -> int:
    blob = f"{line.expense_name}{line.invoice_name}"
    score = 1
    for kws, wt in (
        (["工资", "薪酬", "奖金", "社保", "公积金"], 12),
        (["利息", "手续", "保理", "汇兑"], 11),
        (["房租", "租赁", "物业"], 10),
        (["广宣", "推广", "广告"], 9),
        (["招待"], 7),
        (["差旅"], 6),
        (["研发"], 8),
        (["折旧", "摊销"], 6),
    ):
        if any(k in blob for k in kws):
            score = max(score, wt)
    return score


def _scale_field_to_total(
    lines: list[budget_mod.ExpenseLine],
    field: str,
    target: float,
) -> None:
    """把 lines 上某金额字段等比缩放到 target；全 0 时按权重拆分。"""
    if target <= 0 or not lines:
        return
    cur = sum(float(getattr(l, field, 0) or 0) for l in lines)
    if cur > 1:
        scale = target / cur
        for l in lines:
            v = float(getattr(l, field, 0) or 0)
            if v > 0:
                setattr(l, field, round(v * scale, 2))
        # 尾差
        after = sum(float(getattr(l, field, 0) or 0) for l in lines)
        gap = round(target - after, 2)
        if abs(gap) >= 0.01:
            holders = [l for l in lines if float(getattr(l, field, 0) or 0) > 0]
            if holders:
                setattr(
                    holders[0],
                    field,
                    round(float(getattr(holders[0], field) or 0) + gap, 2),
                )
        return
    # 全空：按权重填到前 N 行
    ranked = sorted(lines, key=_subject_weight, reverse=True)
    pool = ranked[: min(8, len(ranked))]
    w = [_subject_weight(l) for l in pool]
    ws = sum(w) or 1
    for l, wi in zip(pool, w):
        setattr(l, field, round(target * (wi / ws), 2))


def reconcile_plan_to_financials(
    plan: budget_mod.BudgetPlan,
    data: FinancialData,
    *,
    period_expenses: Optional[dict] = None,
) -> list[str]:
    """交付前硬对账：D/I 对齐利润表期间费用；G 按科目贴近「上年×(1+g) 或 本年实际」。

    解决：销售 D 全空、财务 D 虚高拆分但 G 被压扁、AI 填满同名空行等交付错误。
    """
    notes: list[str] = []
    years = sorted(data.years or [])
    latest = years[-1] if years else 0
    prev = years[-2] if len(years) >= 2 else latest
    budget_mod.compute_all(plan)
    c7 = float(plan.top_computed.revenue_growth_rate or 0.0)
    growth = max(0.85, min(1.0 + max(c7, -0.05), 1.25))

    pe = period_expenses if isinstance(period_expenses, dict) else {}
    subject_specs = [
        (
            "销售费用",
            _amount(data, "income", "销售费用", prev),
            _amount(data, "income", "销售费用", latest)
            or float(pe.get("selling_latest") or pe.get("selling_expense") or 0),
        ),
        (
            "管理费用",
            _amount(data, "income", "管理费用", prev)
            + _amount(data, "income", "研发费用", prev),
            (
                _amount(data, "income", "管理费用", latest)
                + _amount(data, "income", "研发费用", latest)
            )
            or (
                float(pe.get("admin_latest") or pe.get("admin_expense") or 0)
                + float(pe.get("rd_latest") or pe.get("rd_expense") or 0)
            ),
        ),
        (
            "财务费用",
            _amount(data, "income", "财务费用", prev),
            _amount(data, "income", "财务费用", latest)
            or float(pe.get("finance_latest") or pe.get("finance_expense") or 0),
        ),
    ]

    for subject, prev_amt, latest_amt in subject_specs:
        lines = [l for l in plan.lines if l.subject == subject]
        if not lines:
            continue
        d_before = sum(float(l.last_year_actual or 0) for l in lines)
        i_before = sum(float(l.actual_amount or 0) for l in lines)
        g_before = sum(float(l.budget_amount or 0) for l in lines)

        # D / I 对齐利润表（相对偏差 >2% 或一侧为 0 时纠正）
        if prev_amt > 0 and (d_before <= 0 or abs(d_before - prev_amt) / prev_amt > 0.02):
            _scale_field_to_total(lines, "last_year_actual", prev_amt)
            notes.append(f"{subject} D 对账 {d_before:,.0f}→{prev_amt:,.0f}")
        if latest_amt > 0 and (i_before <= 0 or abs(i_before - latest_amt) / latest_amt > 0.02):
            _scale_field_to_total(lines, "actual_amount", latest_amt)
            notes.append(f"{subject} I 对账 {i_before:,.0f}→{latest_amt:,.0f}")

        d_sum = sum(float(l.last_year_actual or 0) for l in lines)
        i_sum = sum(float(l.actual_amount or 0) for l in lines)
        # 预算目标：优先 上年×(1+g)，否则本年实际；有 AI 金额时取 max(AI合计, 目标×0.9)
        base_target = 0.0
        if d_sum > 0:
            base_target = d_sum * growth
        elif i_sum > 0:
            base_target = i_sum
        elif latest_amt > 0:
            base_target = latest_amt
        elif prev_amt > 0:
            base_target = prev_amt * growth
        if base_target <= 0:
            continue
        g_sum = sum(float(l.budget_amount or 0) for l in lines)
        # 过低（<目标 80%）或过高（>目标 150% 且有 D/I 锚）则缩放
        if g_sum <= 0:
            _scale_field_to_total(lines, "budget_amount", base_target)
            for l in lines:
                if float(l.budget_amount or 0) > 0 and float(l.last_year_actual or 0) <= 0:
                    if float(l.reference_amount or 0) <= 0:
                        l.reference_amount = float(l.budget_amount)
            notes.append(f"{subject} G 补齐→{base_target:,.0f}")
        elif g_sum < base_target * 0.80 or (base_target > 0 and g_sum > base_target * 1.50):
            _scale_field_to_total(lines, "budget_amount", base_target)
            for l in lines:
                if float(l.last_year_actual or 0) <= 0 and float(l.budget_amount or 0) > 0:
                    l.reference_amount = float(l.budget_amount)
            notes.append(f"{subject} G 对账 {g_before:,.0f}→{base_target:,.0f}")

        # 有 D 的行：参考金额与表公式一致 F≈D×(1+g) 的数值备份
        for l in lines:
            d = float(l.last_year_actual or 0)
            if d > 0:
                l.reference_amount = round(d * (1.0 + c7), 2)
                if float(l.budget_amount or 0) <= 0:
                    l.budget_amount = float(l.reference_amount)

    budget_mod.compute_all(plan)
    notes.append(
        f"利润表对账后 ΣD={plan.last_year_total:,.0f} ΣG={plan.allocated_total:,.0f} "
        f"ΣI={plan.actual_total:,.0f}"
    )
    return notes


def collapse_duplicate_expense_budgets(plan: budget_mod.BudgetPlan) -> list[str]:
    """同科目同费用名称多行：只保留权重最高的一行有 G，其余并入，避免 7 行广宣重复堆数。"""
    from collections import defaultdict

    notes: list[str] = []
    groups: dict[tuple[str, str], list] = defaultdict(list)
    for l in plan.lines:
        key = (l.subject, (l.expense_name or "").strip())
        if float(l.budget_amount or 0) > 0:
            groups[key].append(l)
    n_merged = 0
    for key, rows in groups.items():
        if len(rows) <= 2:
            continue
        # 超过 2 行同名：合并到权重最高的一行（保留最多 2 行有数）
        ranked = sorted(rows, key=_subject_weight, reverse=True)
        keep = ranked[:2]
        drop = ranked[2:]
        pool_g = sum(float(l.budget_amount or 0) for l in drop)
        pool_d = sum(float(l.last_year_actual or 0) for l in drop)
        pool_i = sum(float(l.actual_amount or 0) for l in drop)
        for l in drop:
            l.budget_amount = 0.0
            # D/I 若只在 drop 行，并入 keep
            if float(l.last_year_actual or 0) > 0 and all(
                float(k.last_year_actual or 0) <= 0 for k in keep
            ):
                pass
            l.reference_amount = 0.0 if float(l.last_year_actual or 0) <= 0 else float(
                l.reference_amount or 0
            )
        if pool_g > 0 and keep:
            w = [_subject_weight(l) or 1 for l in keep]
            ws = sum(w) or 1
            for l, wi in zip(keep, w):
                l.budget_amount = round(float(l.budget_amount or 0) + pool_g * (wi / ws), 2)
                if float(l.last_year_actual or 0) <= 0:
                    l.reference_amount = float(l.budget_amount)
            n_merged += len(drop)
        if pool_d > 0 and keep:
            keep[0].last_year_actual = round(float(keep[0].last_year_actual or 0) + pool_d, 2)
            for l in drop:
                l.last_year_actual = 0.0
        if pool_i > 0 and keep:
            keep[0].actual_amount = round(float(keep[0].actual_amount or 0) + pool_i, 2)
            for l in drop:
                l.actual_amount = 0.0
    if n_merged:
        notes.append(f"同名费用行合并：收拢 {n_merged} 行重复预算")
        budget_mod.compute_all(plan)
    return notes


def build_ai_allocation_prompt(data: FinancialData, plan: budget_mod.BudgetPlan) -> tuple[str, str]:
    """构造 DeepSeek 费用分配提示词 (system, user)。"""
    years = sorted(data.years or [])
    latest = years[-1] if years else 0
    prev = years[-2] if len(years) >= 2 else latest
    facts = {
        "years": years,
        "income": {
            acc: {str(y): _amount(data, "income", acc, y) for y in years}
            for acc in ["营业收入", "营业成本", "销售费用", "管理费用", "研发费用", "财务费用"]
        },
        "ledger": {
            acc: {str(y): _amount(data, "ledger", acc, y) for y in years}
            for acc in ["业务招待费", "广告宣传费", "咨询服务费", "福利费", "教育经费", "研发费用", "折旧"]
        },
    }
    catalog = [
        {
            "row": l.row,
            "subject": l.subject,
            "expense_name": l.expense_name,
            "invoice_name": l.invoice_name,
        }
        for l in plan.lines
    ]
    system = (
        "你是中国企业费用预算编制助手。只能使用用户提供的金额事实，"
        "把费用分配到模板行。所有建议必须合法合规。"
        "返回 JSON 对象：{\"lines\":[{\"row\":14,\"last_year_actual\":0,"
        "\"budget_amount\":0,\"actual_amount\":0},...]}。"
        "row 必须是模板已有行号；金额单位元；不要虚构不存在的费用。"
        "预算年费用可按上年×(1+收入增长率)估算；actual 用最新年已发生口径。"
        "只输出 JSON。"
    )
    user = (
        f"企业：{data.company_name} 行业：{data.industry}\n"
        f"预算年≈{latest} 上年≈{prev}\n"
        f"财务事实：{json.dumps(facts, ensure_ascii=False)}\n"
        f"模板行目录（前40行，共{len(catalog)}行）："
        f"{json.dumps(catalog[:40], ensure_ascii=False)}\n"
        "请输出需要填数的行（可只返回非零行，最多 40 条）。"
    )
    return system, user


def finalize_plan(plan: budget_mod.BudgetPlan) -> budget_mod.BudgetPlan:
    """compute + validate，必要时修正非法字段。"""
    budget_mod.compute_all(plan)
    ok, errors = budget_mod.validate_plan(plan)
    if not ok:
        if plan.top_inputs.income_tax_rate not in budget_mod.INCOME_TAX_RATE_CHOICES:
            plan.top_inputs.income_tax_rate = budget_mod.DEFAULT_INCOME_TAX_RATE
        for l in plan.lines:
            l.last_year_actual = max(0.0, float(l.last_year_actual or 0))
            l.budget_amount = max(0.0, float(l.budget_amount or 0))
            l.actual_amount = max(0.0, float(l.actual_amount or 0))
        for field in (
            "budget_revenue", "budget_cost", "last_year_revenue", "last_year_cost"
        ):
            v = getattr(plan.top_inputs, field)
            if v < 0:
                setattr(plan.top_inputs, field, 0.0)
        budget_mod.compute_all(plan)
        ok, errors = budget_mod.validate_plan(plan)
        if not ok:
            raise tpl_mod.TemplateError(
                "预算计划校验未通过：\n  - " + "\n  - ".join(errors)
            )
    if CANONICAL_TEMPLATE.exists():
        plan.source_path = str(CANONICAL_TEMPLATE.resolve())
    return plan


def build_budget_plan_from_session(
    data: FinancialData,
    ocr_texts: Optional[Sequence[str]] = None,
    *,
    top_indicators: Optional[dict] = None,
    line_allocations: Optional[Sequence[dict]] = None,
    period_totals: Optional[dict] = None,
) -> tuple[budget_mod.BudgetPlan, dict]:
    """组装可导出的 BudgetPlan。

    top_indicators / line_allocations / period_totals 由 Web 层 DeepSeek 结果注入；
    未提供时纯规则填充（core 不反向依赖 web_backend）。
    """
    plan = budget_mod.make_empty_plan(
        company_name=data.company_name,
        industry=data.industry or "制造业",
        year=(sorted(data.years)[-1] if data.years else 0),
    )
    meta = {"top_method": "rule", "line_method": "rule", "notes": []}

    if top_indicators:
        # 剥离内部字段
        clean = {k: v for k, v in top_indicators.items() if not str(k).startswith("_")}
        apply_top_indicators(plan, clean)
        meta["top_method"] = "ai"
        meta["notes"].append("AI 识别 · 顶部营收/成本（DeepSeek 多片段）")
        period_totals = period_totals or top_indicators.get("_period")
    else:
        meta["notes"].append(_fill_top_from_data(plan, data))

    # 先 AI 明细（更准），再规则补差额，避免 AI 空行被规则整笔覆盖细节
    if line_allocations:
        n = apply_line_allocations(plan, line_allocations)
        if n > 0:
            meta["line_method"] = "ai"
            meta["notes"].append(f"AI 分配 · 写入 {n} 行费用明细（DeepSeek）")
    meta["notes"].append(_fill_lines_from_data(plan, data))
    if period_totals:
        note = apply_period_totals_to_plan(plan, period_totals)
        if note:
            meta["notes"].append("期间合计补齐 · " + note)

    finalize_plan(plan)
    filled = sum(
        1
        for l in plan.lines
        if (l.last_year_actual or 0) > 0 or (l.budget_amount or 0) > 0 or (l.actual_amount or 0) > 0
    )
    meta["filled_lines"] = filled
    meta["notes"].append(f"非空费用行 {filled}/84")
    return plan, meta


def export_budget_3sheet(
    data: FinancialData,
    path: str,
    *,
    ocr_texts: Optional[Sequence[str]] = None,
    interactive_session=None,
    top_indicators: Optional[dict] = None,
    line_allocations: Optional[Sequence[dict]] = None,
    period_totals: Optional[dict] = None,
    advice_items: Optional[Sequence[dict]] = None,
    ai_rebalance: bool = False,
) -> tuple[str, dict]:
    """写出三 Sheet 预算 Excel，返回 (path, meta)。

    分工（速度 + 准确）：
    - **Web + DeepSeek**：开支预算、未来预期、参考金额/预算金额（advice_items）
    - **本地建模 + Excel 公式**：毛利率、增长率、费用占比 E/H、合计、E7 等
      （不在导出阶段再跑多轮 DeepSeek，显著加快导出）

    advice_items：Web 端编制建议勾选项 → 写入 F/G（无上年不虚构 D）
    ai_rebalance：默认 False；仅无建议且显式打开时才在导出侧二次 AI 占比重算
    """
    plan, meta = build_budget_plan_from_session(
        data,
        ocr_texts=ocr_texts,
        top_indicators=top_indicators,
        line_allocations=line_allocations,
        period_totals=period_totals,
    )
    import importlib

    advice_mod = importlib.import_module("core.CO_budget_advice_WB-CO-TR-20260810")
    period_exp = extract_period_expenses(data)
    meta["period_expenses"] = period_exp
    meta["pipeline"] = "web_ai_amounts + local_model + excel_formulas"

    # 无上年但有预算：补 reference（导出无 D 时 F 写数值）
    for l in plan.lines:
        if float(l.last_year_actual or 0) <= 0 and float(l.budget_amount or 0) > 0:
            if float(l.reference_amount or 0) <= 0:
                l.reference_amount = float(l.budget_amount)

    # ── 1) Web 端 DeepSeek 金额落地（大块业务判断只在此）──
    if advice_items:
        budget_mod.compute_all(plan)
        before = sum(
            1
            for l in plan.lines
            if (l.budget_amount or 0) > 0 or (l.reference_amount or 0) > 0
        )
        advice_mod.apply_advice_to_plan(plan, list(advice_items), optimize_ratios=False)
        selected = sum(
            1
            for it in advice_items
            if isinstance(it, dict) and it.get("selected", True)
        )
        meta["advice_applied"] = True
        meta["advice_selected"] = selected
        after_advice = sum(1 for l in plan.lines if (l.budget_amount or 0) > 0)
        meta["notes"].append(
            f"Web编制建议已写入 · 勾选 {selected} 项 · 有预算行 {before}→{after_advice}"
        )
        meta["amount_source"] = "web_deepseek_advice"
        if selected > 0 and after_advice == 0:
            raise ValueError(
                f"编制建议 {selected} 条未能写入预算表（G 仍全 0）。请重新生成建议后再导出。"
            )
    else:
        meta["amount_source"] = "rules_or_ocr"
        meta["notes"].append("未带 Web 编制建议：使用结构化/规则金额（占比仍由表内公式算）")

    # ── 2) 可选：仅无建议时才二次 AI（默认关闭，加快导出）──
    if ai_rebalance and not advice_items:
        try:
            ai_mod = importlib.import_module("web_backend.CO_ai_WB-CO-TR-20260805160732")
            budget_mod.compute_all(plan)
            c2_rb = float(plan.top_inputs.budget_revenue or 0)
            e7_rb = float(plan.top_computed.expense_budget_cap or 0)
            rb_lines = [
                {
                    "row": l.row,
                    "subject": l.subject,
                    "expense_name": l.expense_name,
                    "invoice_name": l.invoice_name,
                    "last_year_actual": float(l.last_year_actual or 0),
                    "budget_amount": float(l.budget_amount or 0),
                }
                for l in plan.lines
                if float(l.budget_amount or 0) > 0 or float(l.last_year_actual or 0) > 0
            ]
            from . import industry as _ind

            band0 = _ind.get_period_expense_ratio_band(plan.industry)
            rb_items, rb_summary, rb_err = ai_mod.rebalance_expense_ratios(
                {
                    "company_name": plan.company_name,
                    "industry": plan.industry,
                    "budget_revenue": c2_rb,
                    "expense_budget_cap": e7_rb,
                    "revenue_growth_rate": float(plan.top_computed.revenue_growth_rate or 0),
                    "period_expenses": period_exp,
                    "wb_model": {
                        "period_expense_ratio_band": band0,
                        "income_tax_contribution_hub": _ind.get_income_tax_contribution_rate(
                            plan.industry
                        ),
                        "target_fee_total": round(
                            c2_rb * float(band0.get("median") or 0.12), 2
                        ),
                    },
                    "lines": rb_lines,
                    "timeout": 180,
                    "max_passes": 2,
                }
            )
            if rb_items:
                advice_mod.apply_advice_to_plan(plan, rb_items, optimize_ratios=False)
                meta["notes"].append(
                    "导出侧 AI 补算金额（无 Web 建议）"
                    + (f"：{rb_summary[:160]}" if rb_summary else "")
                )
                meta["amount_source"] = "export_deepseek_fallback"
            else:
                meta["notes"].append(f"导出侧 AI 补算未返回：{rb_err or '空'}")
        except Exception as e:
            meta["notes"].append(f"导出侧 AI 补算跳过：{type(e).__name__}")
    else:
        meta["notes"].append(
            "导出跳过 DeepSeek 重算 · 占比/毛利率由表内公式与 WB 本地建模保证"
        )

    # ── 3) 本地建模：WB 顶栏 + 利润表硬对账 + 列补全（确定性，快、准）──
    try:
        from . import industry as ind_mod

        for n in ind_mod.apply_wb_top_rates_to_plan(plan, force=False):
            meta["notes"].append(n)
        budget_mod.compute_all(plan)
        band = ind_mod.get_period_expense_ratio_band(plan.industry)
        meta["wb_fee_band"] = band
        meta["wb_cit_hub"] = ind_mod.get_income_tax_contribution_rate(plan.industry)
        meta["notes"].append(
            f"WB建模：所得税贡献中枢 {meta['wb_cit_hub']*100:.2f}% · "
            f"费用率带 {band.get('min',0):.1%}–{band.get('max',0):.1%} "
            f"（中枢 {band.get('median',0):.1%}）"
        )
    except Exception as e:
        meta["notes"].append(f"WB基准写入跳过：{type(e).__name__}")

    # 关键：D/I/G 与利润表销管财对齐（修「销售 D 空、财务 G 被压扁」）
    for n in reconcile_plan_to_financials(plan, data, period_expenses=period_exp):
        meta["notes"].append(n)
    for n in collapse_duplicate_expense_budgets(plan):
        meta["notes"].append(n)
    for n in advice_mod.ensure_all_required_fields_filled(plan):
        meta["notes"].append(n)
    # 总费用率落在 WB 带内（科目对账后总调，结构已合理）
    for n in advice_mod.align_budget_to_period_level(plan, period_expenses=period_exp):
        meta["notes"].append(n)
    # 再对账一次 G，防止总调破坏科目锚
    for n in reconcile_plan_to_financials(plan, data, period_expenses=period_exp):
        if "G 对账" in n or "G 补齐" in n:
            meta["notes"].append(f"复核·{n}")
    meta["scale_aligned"] = True
    meta["ratio_source"] = "model_precomputed_E_H"  # 占比按 E=D/C6 H=G/C2 预计算写入
    meta["margin_source"] = "model_precomputed_C5_C7_C9"

    qa = advice_mod.finalize_accuracy_qa(plan, period_expenses=period_exp)
    meta["accuracy_qa"] = {
        k: qa[k]
        for k in (
            "ok", "identity_ok", "subject_ok", "c2", "e7", "g_sum",
            "fee_rate", "n_g", "n_f", "missing_f", "subject_sums", "subject_gaps",
        )
        if k in qa
    }
    for n in qa.get("notes") or []:
        meta["notes"].append(n)

    budget_mod.compute_all(plan)
    filled = sum(
        1
        for l in plan.lines
        if (l.last_year_actual or 0) > 0
        or (l.budget_amount or 0) > 0
        or (l.actual_amount or 0) > 0
    )
    meta["filled_lines"] = filled
    g_sum = sum(float(l.budget_amount or 0) for l in plan.lines)
    f_sum = sum(float(l.reference_amount or 0) for l in plan.lines)
    c2 = float(plan.top_inputs.budget_revenue or 0)
    fee_rate = (g_sum / c2) if c2 else 0.0
    meta["advice_budget_sum"] = round(g_sum, 2)
    meta["fee_rate"] = round(fee_rate, 6)
    meta["notes"].append(
        f"定稿：有G {sum(1 for l in plan.lines if (l.budget_amount or 0)>0)}/84 · "
        f"有F {sum(1 for l in plan.lines if (l.reference_amount or 0)>0)}/84 · "
        f"ΣG={g_sum:,.0f} · 模型费用率 {fee_rate:.2%} · "
        f"金额源={meta.get('amount_source')} · 占比=表公式 · "
        f"QA={'通过' if qa.get('ok') else '待核'}"
    )
    missing_f = [
        l.row
        for l in plan.lines
        if float(l.budget_amount or 0) > 0 and float(l.reference_amount or 0) <= 0
    ]
    if missing_f:
        for l in plan.lines:
            if l.row in missing_f:
                l.reference_amount = float(l.budget_amount)
        budget_mod.compute_all(plan)
        meta["notes"].append(f"强制补 F：{len(missing_f)} 行")

    out = tpl_mod.write_template(
        plan, path, session=interactive_session, financial_data=data
    )
    # 月度拆分底稿快照（模块 A 二段式）：行年度金额与写出的 G 列同源；
    # 顶部摘要供第一稿摘要卡展示。附加键不改变既有 meta 契约。
    meta["plan_rows"] = [
        {
            "row": l.row,
            "subject": l.subject,
            "expense_name": l.expense_name,
            "annual": round(float(l.budget_amount or 0), 2),
        }
        for l in plan.lines
    ]
    meta["top_summary"] = {
        "budget_revenue": round(float(plan.top_inputs.budget_revenue or 0), 2),
        "expense_budget_cap": round(float(plan.top_computed.expense_budget_cap or 0), 2),
    }
    meta["path"] = out
    meta["sheets"] = [
        "费用预算表",
        "行业企业所得税贡献率参考",
        "诊断与行动清单",
        "费用合规筹划约束",
    ]
    return out, meta


def append_monthly_sheet(
    xlsx_path: str,
    split_payload: dict,
    out_path: Optional[str] = None,
    mode: Optional[str] = None,
) -> str:
    """在第一稿工作簿末尾追加「月度执行计划」Sheet，返回新文件路径。

    - 不改任何现有 Sheet（read_template 只校验指定 Sheet，天然兼容追加）；
    - Q 列合计用公式 =SUM(E:P) 可复算，R 列校验 =Q−D 应全 0（非 0 条件标红）；
    - 同名 Sheet 重复调用幂等覆盖；
    - 样式沿用纸墨台账：表头加粗 + 发丝边框，无花哨填充。
    """
    from openpyxl import load_workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    src = Path(xlsx_path)
    if not src.exists():
        raise FileNotFoundError(f"第一稿文件不存在：{xlsx_path}")
    if out_path is None:
        out_path = str(src.with_name(f"{src.stem}（含月度拆分）.xlsx"))
    mode = mode or str((split_payload or {}).get("mode") or "rule")

    wb = load_workbook(str(src))
    sheet_name = "月度执行计划"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    headers = (
        ["行号", "科目名称", "费用项目", "年度预算金额"]
        + [f"{i}月" for i in range(1, 13)]
        + ["月度合计", "校验（合计−年度）"]
    )
    hair = Side(style="hair", color="9C8B6A")
    border = Border(left=hair, right=hair, top=hair, bottom=hair)
    header_font = Font(bold=True)
    paper_fill = PatternFill("solid", fgColor="F3EDE2")
    center = Alignment(horizontal="center", vertical="center")

    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.border = border
        cell.alignment = center
        cell.fill = paper_fill

    rows = sorted(
        (split_payload or {}).get("matrix") or [],
        key=lambda r: int(r.get("row") or 0),
    )
    money_format = "#,##0.00"
    first_data = 2
    for idx, item in enumerate(rows):
        r = first_data + idx
        ws.cell(row=r, column=1, value=int(item.get("row") or 0))
        ws.cell(row=r, column=2, value=str(item.get("subject") or ""))
        ws.cell(row=r, column=3, value=str(item.get("expense_name") or ""))
        ws.cell(row=r, column=4, value=float(item.get("annual") or 0)).number_format = money_format
        for m in range(12):
            months = item.get("months") or []
            value = float(months[m]) if m < len(months) else 0.0
            ws.cell(row=r, column=5 + m, value=value).number_format = money_format
        ws.cell(row=r, column=17, value=f"=SUM(E{r}:P{r})").number_format = money_format
        ws.cell(row=r, column=18, value=f"=Q{r}-D{r}").number_format = money_format
        for col in range(1, 19):
            c = ws.cell(row=r, column=col)
            c.border = border
            if col in (1, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18):
                c.alignment = center
    last_data = first_data + len(rows) - 1

    # 月度总计行 + 说明行
    total_row = last_data + 1
    ws.cell(row=total_row, column=1, value="合计")
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    for col in range(5, 17):
        letter = ws.cell(row=1, column=col).coordinate.rstrip("0123456789")
        ws.cell(
            row=total_row, column=col,
            value=f"=SUM({letter}{first_data}:{letter}{last_data})",
        ).number_format = money_format
    ws.cell(row=total_row, column=17, value=f"=SUM(Q{first_data}:Q{last_data})").number_format = money_format
    ws.cell(row=total_row, column=18, value=f"=Q{total_row}-SUM(D{first_data}:D{last_data})").number_format = money_format
    for col in range(1, 19):
        c = ws.cell(row=total_row, column=col)
        c.border = border
        c.font = Font(bold=True)
        if col >= 4:
            c.alignment = center

    checks = (split_payload or {}).get("checks") or {}
    note = (
        f"拆分方式：{'AI 权重（金额由确定性引擎计算）' if mode == 'ai' else '规则默认拆分'}；"
        f"生成时间：{(split_payload or {}).get('generated_at') or ''}；"
        f"恒等校验：逐行 12 个月合计 = 年度预算（失败行 {int(checks.get('row_failures') or 0)}，"
        f"总偏差 {float(checks.get('total_gap') or 0):,.2f} 元）；校验列（本列右一）应全为 0；金额单位：元。"
    )
    note_cell = ws.cell(row=total_row + 2, column=1, value=note)
    note_cell.font = Font(size=9, color="5A4E3D")

    # 校验列非 0 标红（数据行 + 总计行）
    ws.conditional_formatting.add(
        f"R{first_data}:R{total_row}",
        CellIsRule(operator="notEqual", formula=["0"], font=Font(color="B00020", bold=True)),
    )

    widths = {1: 6, 2: 12, 3: 18, 4: 14}
    for col in range(5, 19):
        widths[col] = 11
    for col, width in widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col).coordinate.rstrip("0123456789")].width = width
    ws.freeze_panes = "A2"

    wb.save(out_path)
    return out_path
