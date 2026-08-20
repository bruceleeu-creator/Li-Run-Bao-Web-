"""协同看板 · 跟踪进度表 Excel 模板（openpyxl 生成/解析/导出）。

列契约（表头严格匹配前 8 列）：
A 任务名称*｜B 详细说明｜C 负责人（用户名）｜D 状态｜E 截止日期｜F 优先级｜G 月份归属｜H 备注
导出另含 I 创建人 / J 完成时间（只读）；删 I/J 后可原样导回（往返一致）。
导入上限 500 行/次；负责人不存在 → 任务照常导入并标「待分配」。
"""
from __future__ import annotations

import importlib
import io
import re
from datetime import date, datetime
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl.styles import Font
from pydantic import BaseModel

db = importlib.import_module("board_backend.CO_db_WB-CO-TR-20260820")
rooms = importlib.import_module("board_backend.CO_rooms_WB-CO-TR-20260820")

router = APIRouter(prefix="/api", tags=["template"])

HEADERS = ["任务名称*", "详细说明", "负责人（用户名）", "状态", "截止日期", "优先级", "月份归属", "备注"]
EXPORT_HEADERS = HEADERS + ["创建人", "完成时间"]
IMPORT_ROW_LIMIT = 500

_STATUS_MAP = {"待办": "todo", "进行中": "doing", "已完成": "done",
               "todo": "todo", "doing": "doing", "done": "done"}
_PRIORITY_MAP = {"高": "high", "中": "mid", "低": "low",
                 "high": "high", "mid": "mid", "low": "low"}


def _write_headers(ws, headers) -> None:
    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name).font = Font(bold=True)


def _attachment(name: str) -> str:
    """中文文件名用 RFC 5987 编码（HTTP 头仅允许 latin-1）。"""
    return f"attachment; filename*=UTF-8''{quote(name)}"


@router.get("/template.xlsx")
def download_template() -> StreamingResponse:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "跟踪进度表"
    _write_headers(ws, HEADERS)
    ws.cell(row=2, column=1, value="示例：完成 3 月广宣投放核销（可删除本行）")
    ws.cell(row=2, column=4, value="待办")
    ws.cell(row=2, column=5, value=date.today().isoformat())
    ws.cell(row=2, column=6, value="中")
    ws.cell(row=2, column=7, value=f"{date.today().strftime('%Y-%m')}")
    for col, width in zip(range(1, 9), (36, 40, 16, 10, 12, 8, 12, 24)):
        ws.column_dimensions[ws.cell(row=1, column=col).coordinate.rstrip("0123456789")].width = width
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": _attachment("跟踪进度表模板.xlsx")},
    )


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip()


def _parse_due(v) -> str | None:
    s = _cell_str(v)
    if not s:
        return None
    m = re.match(r"^(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3))).isoformat()
        except ValueError:
            return None
    return None


@router.post("/rooms/{room_id}/import.xlsx")
async def import_template(room_id: int, file: UploadFile, ctx: dict = Depends(rooms.require_member)) -> dict:
    from openpyxl import load_workbook

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="无法解析文件，请使用下载的官方模板填写")
    ws = wb.active
    header_row = [_cell_str(ws.cell(1, c).value) for c in range(1, 9)]
    if header_row != HEADERS:
        raise HTTPException(status_code=400, detail="表头与模板不一致，请下载官方模板后填写再导入")

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if len(rows) > IMPORT_ROW_LIMIT:
        raise HTTPException(status_code=400, detail=f"单次最多导入 {IMPORT_ROW_LIMIT} 行")

    user_id = ctx["user"]["id"]
    results: list[dict] = []
    with db.conn() as c:
        members = c.execute(
            """
            SELECT u.id, u.username FROM room_members m JOIN users u ON u.id = m.user_id
            WHERE m.room_id = %s
            """,
            (room_id,),
        ).fetchall()
        by_name = {m["username"]: m["id"] for m in members}
        for idx, row in enumerate(rows, start=2):
            cells = list(row[:8]) + [""] * max(0, 8 - len(row))
            title = _cell_str(cells[0])[:100]
            if not title:
                results.append({"row": idx, "status": "skipped", "reason": "任务名称为空"})
                continue
            assignee_name = _cell_str(cells[2])
            assignee_id = by_name.get(assignee_name) if assignee_name else None
            status_raw = _cell_str(cells[3])
            status = _STATUS_MAP.get(status_raw, "todo")
            priority_raw = _cell_str(cells[5])
            priority = _PRIORITY_MAP.get(priority_raw, "mid")
            month_raw = _cell_str(cells[6])
            month_tag = month_raw if re.match(r"^\d{4}-(0[1-9]|1[0-2])$", month_raw) else None
            note_bits = []
            if assignee_name and assignee_id is None:
                note_bits.append("负责人不存在→待分配")
            if status_raw and status_raw not in _STATUS_MAP:
                note_bits.append(f"状态「{status_raw}」非法→待办")
            if priority_raw and priority_raw not in _PRIORITY_MAP:
                note_bits.append(f"优先级「{priority_raw}」非法→中")
            inserted = c.execute(
                """
                INSERT INTO tasks (room_id, title, detail, status, assignee_id, creator_id,
                                   due_date, priority, month_tag)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id
                """,
                (room_id, title, _cell_str(cells[1]), status, assignee_id, user_id,
                 _parse_due(cells[4]), priority, month_tag),
            ).fetchone()
            c.execute(
                "INSERT INTO task_events (room_id, task_id, user_id, action, payload) VALUES (%s, %s, %s, 'import', %s)",
                (room_id, inserted["id"], user_id,
                 __import__("json").dumps({"row": idx, "title": title}, ensure_ascii=False)),
            )
            results.append({
                "row": idx, "status": "created", "task_id": inserted["id"],
                "reason": "；".join(note_bits) if note_bits else "",
            })
        c.execute("UPDATE rooms SET version = version + 1 WHERE id = %s", (room_id,))
    created = sum(1 for r in results if r["status"] == "created")
    skipped = sum(1 for r in results if r["status"] == "skipped")
    return {"created": created, "skipped": skipped, "rows": results}


@router.get("/rooms/{room_id}/export.xlsx")
def export_room(room_id: int, ctx: dict = Depends(rooms.require_member)) -> StreamingResponse:
    from openpyxl import Workbook

    with db.conn() as c:
        tasks = c.execute(
            "SELECT * FROM tasks WHERE room_id = %s ORDER BY sort_no, id", (room_id,)
        ).fetchall()
        # 含历史创建人/负责人（成员被移出后导出仍显示其快照名）
        members = c.execute(
            """
            SELECT DISTINCT u.id, u.username, u.display_name
            FROM users u
            WHERE u.id IN (
              SELECT user_id FROM room_members WHERE room_id = %s
              UNION
              SELECT creator_id FROM tasks WHERE room_id = %s
              UNION
              SELECT assignee_id FROM tasks WHERE room_id = %s AND assignee_id IS NOT NULL
            )
            """,
            (room_id, room_id, room_id),
        ).fetchall()
    by_id = {m["id"]: m for m in members}
    status_cn = {"todo": "待办", "doing": "进行中", "done": "已完成"}
    prio_cn = {"high": "高", "mid": "中", "low": "低"}

    wb = Workbook()
    ws = wb.active
    ws.title = "跟踪进度表"
    _write_headers(ws, EXPORT_HEADERS)
    for r_idx, t in enumerate(tasks, start=2):
        creator = by_id.get(t["creator_id"])
        ws.cell(row=r_idx, column=1, value=t["title"])
        ws.cell(row=r_idx, column=2, value=t["detail"])
        assignee = by_id.get(t["assignee_id"]) if t["assignee_id"] else None
        ws.cell(row=r_idx, column=3, value=assignee["username"] if assignee else "")
        ws.cell(row=r_idx, column=4, value=status_cn.get(t["status"], t["status"]))
        ws.cell(row=r_idx, column=5, value=t["due_date"].isoformat() if t["due_date"] else None)
        ws.cell(row=r_idx, column=6, value=prio_cn.get(t["priority"], t["priority"]))
        ws.cell(row=r_idx, column=7, value=t["month_tag"])
        ws.cell(row=r_idx, column=8, value="")
        ws.cell(row=r_idx, column=9,
                value=creator["display_name"] if creator else "")
        ws.cell(row=r_idx, column=10,
                value=t["completed_at"].strftime("%Y-%m-%d %H:%M") if t["completed_at"] else None)
    for col, width in zip(range(1, 11), (36, 40, 16, 10, 12, 8, 12, 24, 14, 18)):
        ws.column_dimensions[ws.cell(row=1, column=col).coordinate.rstrip("0123456789")].width = width
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    name = (ctx["room"]["name"] or "看板")[:20]
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": _attachment(f"{name}-跟踪进度表.xlsx")},
    )
