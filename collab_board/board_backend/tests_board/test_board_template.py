"""模板：下载 / 表头校验 / 导入逐行反馈 / 导出往返一致。"""
from __future__ import annotations

import io

from openpyxl import load_workbook

from tests_board.conftest import auth_headers, make_room, register


def _setup(client):
    boss = register(client, "tpl_boss")
    room = make_room(client, boss["token"], "模板测试")
    staff = register(client, "tpl_staff")
    client.post(f"/api/rooms/{room['id']}/join",
                json={"invite_code": room["invite_code"]},
                headers=auth_headers(staff["token"]))
    return boss, staff, room


def _fill_template(content: bytes) -> bytes:
    """在官方模板基础上填 5 行（含 1 行负责人乱填、1 行状态乱填）。"""
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    ws.cell(row=2, column=1, value="核销 3 月广宣发票")
    ws.cell(row=2, column=3, value="tpl_staff")
    ws.cell(row=2, column=4, value="进行中")
    ws.cell(row=2, column=5, value="2026-09-30")
    ws.cell(row=2, column=6, value="高")
    ws.cell(row=2, column=7, value="2026-09")
    ws.cell(row=3, column=1, value="负责人乱填的任务")
    ws.cell(row=3, column=3, value="no_such_user")
    ws.cell(row=4, column=1, value="状态乱填的任务")
    ws.cell(row=4, column=4, value="莫名其状")
    ws.cell(row=5, column=1, value="正常待办")
    ws.cell(row=6, column=2, value="只有说明没有名称（应跳过）")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_template_download_and_import(client):
    boss, staff, room = _setup(client)
    h = auth_headers(boss["token"])

    r = client.get("/api/template.xlsx")
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    headers = [wb.active.cell(1, c).value for c in range(1, 9)]
    assert headers[0] == "任务名称*" and headers[7] == "备注"

    filled = _fill_template(r.content)
    r = client.post(f"/api/rooms/{room['id']}/import.xlsx",
                    files={"file": ("跟踪进度表.xlsx", io.BytesIO(filled),
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                    headers=h)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["created"] == 4  # 5 行中 1 行无名称被跳过
    assert body["skipped"] == 1
    by_row = {x["row"]: x for x in body["rows"]}
    assert "待分配" in by_row[3]["reason"]      # 负责人不存在
    assert "非法" in by_row[4]["reason"]         # 状态乱填回落待办

    board = client.get(f"/api/rooms/{room['id']}/board", headers=h).json()
    tasks = {t["title"]: t for t in board["tasks"]}
    assert tasks["负责人乱填的任务"]["assignee_id"] is None
    assert tasks["状态乱填的任务"]["status"] == "todo"
    assert tasks["核销 3 月广宣发票"]["status"] == "doing"
    assert tasks["核销 3 月广宣发票"]["priority"] == "high"


def test_import_bad_header_rejected(client):
    boss, _, room = _setup(client)
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.cell(1, 1, value="随手写的表头")
    buf = io.BytesIO()
    wb.save(buf)
    r = client.post(f"/api/rooms/{room['id']}/import.xlsx",
                    files={"file": ("bad.xlsx", io.BytesIO(buf.getvalue()),
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                    headers=auth_headers(boss["token"]))
    assert r.status_code == 400
    assert "官方模板" in r.json()["detail"]


def test_export_roundtrip(client):
    boss, staff, room = _setup(client)
    h = auth_headers(boss["token"])
    client.post(f"/api/rooms/{room['id']}/tasks",
                json={"title": "往返任务A", "assignee_id": None}, headers=h)
    client.post(f"/api/rooms/{room['id']}/tasks",
                json={"title": "往返任务B", "detail": "备注信息"}, headers=auth_headers(staff["token"]))

    r = client.get(f"/api/rooms/{room['id']}/export.xlsx", headers=h)
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert ws.cell(1, 9).value == "创建人" and ws.cell(1, 10).value == "完成时间"
    rows = [[ws.cell(rr, cc).value for cc in range(1, 9)] for rr in range(2, ws.max_row + 1)]
    assert len(rows) == 2

    # 删 I/J 两列后原样导回（往返一致）
    wb2 = load_workbook(io.BytesIO(r.content))
    ws2 = wb2.active
    ws2.delete_cols(9, 2)
    buf = io.BytesIO()
    wb2.save(buf)
    r = client.post(f"/api/rooms/{room['id']}/import.xlsx",
                    files={"file": ("reimport.xlsx", io.BytesIO(buf.getvalue()),
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                    headers=h)
    assert r.status_code == 200
    assert r.json()["created"] == 2  # 条数相符
    board = client.get(f"/api/rooms/{room['id']}/board", headers=h).json()
    titles = [t["title"] for t in board["tasks"]]
    assert titles.count("往返任务A") == 2 and titles.count("往返任务B") == 2  # 原有 + 导回
